"""
National Benchmark Survey — May 2025
Weighting Script (Wisconsin methodology — single-CSV edition)

WHAT THIS DOES:
  • Loads a single cleaned Pollfish CSV (Pollfish_Cleaned_Trimmed_Columns.csv)
  • Auto-detects every Q2–Q25 column via fuzzy prefix scan (Wisconsin methodology)
  • Applies iterative raking to 7 dimensions:
      1. Age × Gender joint cells (8 cells — each age group gets correct M/F %)
      2. Race × Education (White split into White No College / White College)
      3. Education 4-way
      4. Region (8 national Census regions — STATE_TO_REGION)
      5. Gender × Education joint cells
      6. Vote2024 INCLUDING "Did not vote" (FEC + CPS non-voter share)
      7. Vote history bucket
  • Entropy Balancing (Method 5) before raking rounds
  • DEFF-informed weight capping (4 raking rounds, cap set after round 1)
  • Response Propensity Scores (Method 8)
  • Two-stage 2024 recall calibration (FEC voters + CPS non-voter anchor)
  • Likely Voter model (Q3/Q4/Q5 propensity × Q2 vote history)
  • Full tabbook Excel export with net rows (Approve/Disapprove/Net,
    horse-race R+/D+, More/Less-Likely) — Wisconsin design system
  • Q7 ranking block: mean rank + % Ranked #1 sub-section
  • Weight Diagnostics sheet (DEFF, Kish, MoE, covariate balance, ICC)
  • Bootstrap CI sheet (Method 6)
  • Electorate Composition sheet

USAGE:
  python usmay.py [/path/to/Pollfish_Cleaned_Trimmed_Columns.csv]
  Default: looks for CSV in current working directory.
"""

import sys, io, os, time as _time, textwrap, warnings
from copy import deepcopy
import pandas as pd
import numpy as np
from scipy.optimize import minimize
from sklearn.linear_model import LogisticRegression

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 0) GLOBAL HELPERS — STATE → 8-REGION MAP (from usweightsmarch23.py)
# ==============================================================================

STATE_TO_REGION = {
    "Alabama": "Appalachia / South Interior", "Alaska": "West / Mountain / Pacific",
    "Arizona": "Southwest", "Arkansas": "Lower Midwest / Plains",
    "California": "West / Mountain / Pacific", "Colorado": "West / Mountain / Pacific",
    "Connecticut": "New England", "Delaware": "Mid-Atlantic",
    "Florida": "Southeast Atlantic", "Georgia": "Southeast Atlantic",
    "Hawaii": "West / Mountain / Pacific", "Idaho": "West / Mountain / Pacific",
    "Illinois": "Great Lakes", "Indiana": "Great Lakes",
    "Iowa": "Lower Midwest / Plains", "Kansas": "Lower Midwest / Plains",
    "Kentucky": "Appalachia / South Interior", "Louisiana": "Appalachia / South Interior",
    "Maine": "New England", "Maryland": "Mid-Atlantic",
    "Massachusetts": "New England", "Michigan": "Great Lakes",
    "Minnesota": "Great Lakes", "Mississippi": "Appalachia / South Interior",
    "Missouri": "Lower Midwest / Plains", "Montana": "West / Mountain / Pacific",
    "Nebraska": "Lower Midwest / Plains", "Nevada": "West / Mountain / Pacific",
    "New Hampshire": "New England", "New Jersey": "Mid-Atlantic",
    "New Mexico": "Southwest", "New York": "Mid-Atlantic",
    "North Carolina": "Southeast Atlantic", "North Dakota": "Lower Midwest / Plains",
    "Ohio": "Great Lakes", "Oklahoma": "Southwest",
    "Oregon": "West / Mountain / Pacific", "Pennsylvania": "Mid-Atlantic",
    "Rhode Island": "New England", "South Carolina": "Southeast Atlantic",
    "South Dakota": "Lower Midwest / Plains", "Tennessee": "Appalachia / South Interior",
    "Texas": "Southwest", "Utah": "West / Mountain / Pacific",
    "Vermont": "New England", "Virginia": "Southeast Atlantic",
    "Washington": "West / Mountain / Pacific", "West Virginia": "Appalachia / South Interior",
    "Wisconsin": "Great Lakes", "Wyoming": "West / Mountain / Pacific",
    "District of Columbia": "Mid-Atlantic", "DC": "Mid-Atlantic",
}

NATIONAL_REGIONS = [
    "New England", "Mid-Atlantic", "Southeast Atlantic", "Appalachia / South Interior",
    "Great Lakes", "Lower Midwest / Plains", "Southwest", "West / Mountain / Pacific",
]

def normalize_state_to_region(state_raw: str) -> str:
    """Map Pollfish state field to 8-region bucket. Handles 'Illinois (US-IL)' format."""
    s = str(state_raw).strip()
    if "(" in s:
        s = s.split("(")[0].strip()
    if "," in s:
        s = [p.strip() for p in s.split(",")][-1]
    if s in STATE_TO_REGION:
        return STATE_TO_REGION[s]
    for state, region in STATE_TO_REGION.items():
        if state.lower() in s.lower() or s.lower() in state.lower():
            return region
    print(f"  [WARNING] normalize_state_to_region: unrecognized state {state_raw!r} — defaulting to 'Southeast Atlantic'")
    return "Southeast Atlantic"  # fallback

# ==============================================================================
# CSV INPUT (single cleaned file)
# ==============================================================================

INPUT_FILE  = "Pollfish_Cleaned_Trimmed_Columns.csv"

# ==============================================================================
# NORMALIZATION HELPERS (Wisconsin methodology)
# ==============================================================================

def normalize_gender(val: str) -> str:
    v = str(val).strip().lower()
    if v in ("male", "m", "man"): return "Male"
    if v in ("female", "f", "woman"): return "Female"
    return "Other"

def normalize_race_edu(race_raw: str, edu_raw: str, ethnicity_raw: str = "") -> str:
    """
    Split White into White No College / White College.
    Other groups kept as named categories.
    Mirrors Wisconsin normalize_race_edu exactly.
    """
    eth = str(ethnicity_raw).strip().lower()
    if "hispanic" in eth or "latino" in eth:
        return "Hispanic"
    race = str(race_raw).strip().lower()
    if "black" in race or "african" in race:
        return "Black"
    asian_terms = ["asian","chinese","japanese","korean","filipino",
                   "vietnamese","indian","pacific","pacific islander"]
    if any(t in race for t in asian_terms):
        return "Asian / Other"
    if "white" in race or "caucasian" in race:
        edu = str(edu_raw).strip().lower()
        college_terms = ["bachelor","master","doctorate","phd","postgrad",
                         "graduate degree","professional degree"]
        if any(t in edu for t in college_terms):
            return "White College"
        return "White No College"
    return "Asian / Other"

def normalize_education_binary(edu_raw: str) -> str:
    edu = str(edu_raw).strip().lower()
    if any(t in edu for t in ["bachelor","master","doctorate","phd",
                                "postgrad","graduate degree"]):
        return "College"
    return "No College"

def normalize_gender_edu(gender_val: str, edu_val: str) -> str:
    """
    Cross-tabulated Gender × Education cell for raking.
    Keys match GenderEdu benchmark: 'Male_College', 'Male_No College',
    'Female_College', 'Female_No College'.
    """
    gender = normalize_gender(gender_val)
    edu    = normalize_education_binary(edu_val)
    return f"{gender}_{edu}"

def normalize_education_4way(edu_raw: str) -> str:
    edu = str(edu_raw).strip().lower()
    if any(t in edu for t in ["postgrad","master","doctorate","phd",
                               "professional degree","graduate degree"]):
        return "Postgraduate study"
    if any(t in edu for t in ["bachelor","college grad","4-year","university grad","undergrad"]):
        return "College graduate"
    if any(t in edu for t in ["some college","associate","assoc","vocational","technical","2-year"]):
        return "Some college/assoc. degree"
    return "High school or less"

def party_from_q8(val: str) -> str:
    s = str(val).strip().lower()
    if "lean republican independent" in s or "lean democratic independent" in s:
        return "Independent"
    if "independent" in s or "none of these" in s or "no clear" in s:
        return "Independent"
    if "republican" in s: return "Republican"
    if "democrat" in s:   return "Democrat"
    return "Independent"

def vote2024_bucket(val: str) -> str:
    v = str(val).strip()
    if v in ("Donald Trump","Kamala Harris","Third party","Did not vote"):
        return v
    low = v.lower()
    if "trump"   in low: return "Donald Trump"
    if "harris"  in low: return "Kamala Harris"
    if "third"   in low: return "Third party"
    if "did not" in low or "didn't" in low: return "Did not vote"
    return np.nan

def clean_text_series(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
    return s.replace({"nan": np.nan, "": np.nan, "None": np.nan})

def normalize_income(val) -> str:
    """Map Pollfish granular income bands to 7 reporting buckets."""
    v = str(val).strip()
    _UNDER_25 = [
        "Less than $5,000", "$5,000 to $9,999", "$10,000 to $14,999",
        "$15,000 to $19,999", "$20,000 to $24,999",
    ]
    _25_50 = [
        "$25,000 to $29,999", "$30,000 to $34,999", "$35,000 to $39,999",
        "$40,000 to $44,999", "$45,000 to $49,999",
    ]
    _50_75 = [
        "$50,000 to $54,999", "$55,000 to $59,999", "$60,000 to $64,999",
        "$65,000 to $69,999", "$70,000 to $74,999",
    ]
    _75_100 = [
        "$75,000 to $79,999", "$80,000 to $84,999", "$85,000 to $89,999",
        "$90,000 to $94,999", "$95,000 to $99,999",
    ]
    _100_150 = ["$100,000 to $124,999", "$125,000 to $149,999"]
    _150_200 = ["$150,000 to $174,999", "$175,000 to $199,999"]
    _200_PLUS = [
        "$200,000 to $249,999", "$250,000 to $499,999",
        "$500,000 to $999,999", "$1 million +",
    ]
    if v in _UNDER_25:  return "$0–$25k"
    if v in _25_50:     return "$25–$50k"
    if v in _50_75:     return "$50–$75k"
    if v in _75_100:    return "$75–$100k"
    if v in _100_150:   return "$100–$150k"
    if v in _150_200:   return "$150–$200k"
    if v in _200_PLUS:  return "$200k+"
    return np.nan  # "Prefer not to answer" and unknown → excluded

# ==============================================================================
# 1) LOAD & COMBINE 4 CSVs
# ==============================================================================


# ==============================================================================
# 2) AUTO-DETECT & RENAME COLUMNS (Wisconsin fuzzy-rename engine)
#    Scans every column for Pollfish question prefixes — first match wins.
# ==============================================================================

_QUESTION_PREFIX_MAP = {
    # Vote history multi-select (Pollfish exports one column per checkbox)
    "Q2_Voted_2024":    ["election years did you vote in at least once?_2024",
                         "election years_2024 - Presidential"],
    "Q2_Voted_2022":    ["election years_2022", "election years did you vote_2022"],
    "Q2_Voted_2020":    ["election years_2020", "election years did you vote_2020"],
    "Q2_Voted_2018":    ["election years_2018", "election years did you vote_2018"],
    "Q2_Voted_2016":    ["election years_2016", "election years did you vote_2016"],
    "Q2_Voted_2014":    ["election years_2014", "election years did you vote_2014"],
    "Q2_Never_Voted":   ["I have never voted","never voted","wasn't eligible",
                          "not eligible during any","did not vote in any of these"],
    # Q3-Q6
    "Q3_VoteIntent":    ["intention and motivation to vote","How would you describe your intention"],
    "Q4_BallotMethod":  ["How do you plan to cast your ballot","plan to cast your ballot"],
    "Q5_SocialVote":    ["5.10 people you are closest","5–10 people you are closest",
                          "people you are closest to, how many"],
    "Q6_2024Vote":      ["Who did you vote for in the 2024 Presidential",
                          "2024 Presidential Election"],
    # Q8-Q11
    "Q8_PoliticalOutlook":  ["best describes your general political outlook",
                              "general political outlook"],
    "Q9_PartyID":           ["Which political party do you identify",
                              "political party do you identify"],
    "Q10_Groyper":          ["Do you identify as a Groyper",
                             "identify as a Groyper",
                             "Groyper, an individual who supports Nick Fuentes",
                             "Groyper","Nick Fuentes"],
    "Q11_RightTrack":       ["direction of the country on the right track",
                              "right track or the wrong track"],
    # Q13-Q14
    "Q13_GenericBallot":    ["2026 Midterm elections were held today",
                              "If the 2026 Midterm"],
    "Q14_TrumpApprove":     ["approve or disapprove of Donald J. Trump",
                              "Trump.s performance as President",
                              "Trump's performance as President"],
    # Q15 open-ended → skip
    # Q18-Q25
    "Q18_TrumpConservatism":["describe President Trump.s actions and policies",
                              "describe President Trump's actions and policies",
                              "Far too conservative"],
    "Q19_ForeignPolicy":    ["whose interests do President Trump.s foreign policy",
                              "whose interests do President Trump's foreign policy",
                              "foreign policy decisions primarily serve"],
    "Q20_HHExpenses":       ["difficult has it been for you to pay for your usual",
                              "household expenses"],
    "Q21_MassDeportation":  ["support or oppose the mass deportation",
                              "mass deportation of all illegal"],
    "Q22_GenPref":          ["generation do you most prefer that candidate",
                              "voting for a candidate for public office, which generation"],
    "Q23_PartnerQuality":   ["single most important quality to you when considering",
                              "long-term partner"],
    "Q24_IsraelPAC":        ["candidate for public office accepted donations from",
                              "PAC that supports Israel"],
    "Q25_CharlieKirk":      ["responsible for the assassination of Charlie Kirk",
                              "Charlie Kirk at Utah Valley"],
    # Demographics (Pollfish appended columns)
    "DEMO_Gender":          ["TPSI Gender","gender","What is your gender"],
    "DEMO_Race":            ["TPSI Race","race","racial or ethnic heritage",
                              "What is your racial"],
    "DEMO_Ethnicity":       ["TPSI Ethnicity","ethnicity","Hispanic","Latino"],
    "DEMO_Education":       ["TPSI Education","education","level of education",
                              "best describes your level of education"],
    "DEMO_State":           ["TPSI State","TPSI Region","state","State"],
    "DEMO_Income":          ["TPSI Income","income","household income"],
}

MATRIX_PREFIXES = {
    "Q12_2028_Matchups": ["If the 2028 Presidential Election were held today",
                           "2028 Presidential Election were held today"],
    "Q16_IssueHandling": ["Based on President Trump",
                           "rate his approval rating on his handling"],
    "Q17_IndivApproval": ["Do you approve of disapprove of the following",
                           "approve of disapprove of the following individuals"],
}

Q7_RANKING_ISSUES = [
    "Economy, Jobs & Cost of Living",
    "Immigration & Border Security",
    "Crime, Public Safety & Policing",
    "Foreign Policy & National Security",
    "Healthcare, Social Security & Medicare",
    "Education, Housing & Family Issues",
    "Energy, Climate & the Environment",
    "Guns & Second Amendment Rights",
    "Civil Rights, Personal Freedoms & Social Issues",
    "Political Corruption, Lobbying & Money in Politics",
]

def auto_rename_columns(df_raw):
    """
    Fuzzy rename engine from Wisconsin: scans all column names for
    prefix candidates. Prints every rename for auditability.
    """
    df = df_raw.copy()
    renames = {}
    print("\n  AUTO-DETECTING COLUMNS:")
    for target, candidates in _QUESTION_PREFIX_MAP.items():
        if target in df.columns:
            continue
        for cand in candidates:
            matches = [c for c in df.columns
                       if isinstance(c, str) and cand.lower() in c.lower()]
            if matches:
                best = min(matches, key=len)
                if best not in renames.values():
                    renames[best] = target
                    print(f"    Renamed: {best[:65]!r}  →  {target}")
                break
        else:
            pass  # silent — column may just not exist in this survey

    df = df.rename(columns=renames)

    # Resolve matrix prefixes (find actual column prefix present in df)
    resolved_matrix = {}
    for key, candidates in MATRIX_PREFIXES.items():
        for cand in candidates:
            hits = [c for c in df.columns
                    if isinstance(c, str) and cand.lower() in c.lower()]
            if hits:
                resolved_matrix[key] = cand
                print(f"    Matrix prefix [{key}]: {cand!r} ({len(hits)} cols matched)")
                break
        else:
            resolved_matrix[key] = None

    # Q7 ranking columns
    Q7_RANK_COLS = {}
    for issue in Q7_RANKING_ISSUES:
        matches = [c for c in df.columns
                   if isinstance(c, str) and "Q7" in c and issue[:15].lower() in c.lower()]
        if not matches:
            matches = [c for c in df.columns
                       if isinstance(c, str) and issue[:20].lower() in c.lower()
                       and "rank" in c.lower()]
        if matches:
            Q7_RANK_COLS[issue] = matches[0]
    print(f"    Q7 ranking: {len(Q7_RANK_COLS)}/{len(Q7_RANKING_ISSUES)} issues matched")

    # Q2 vote history multi-select columns
    Q2_HISTORY_COLS = {}
    Q2_ELECTIONS = ["2024","2022","2020","2018","2016","2014"]
    for yr in Q2_ELECTIONS:
        key = f"Q2_Voted_{yr}"
        if key in df.columns:
            Q2_HISTORY_COLS[yr] = key
        else:
            fallback = [c for c in df.columns
                        if isinstance(c, str) and "Q2" in c and yr in c
                        and "election" in c.lower()]
            if fallback:
                Q2_HISTORY_COLS[yr] = fallback[0]

    return df, resolved_matrix, Q7_RANK_COLS, Q2_HISTORY_COLS

# ==============================================================================
# 3) NATIONAL WEIGHTING BENCHMARKS
# ==============================================================================
# Sources clearly marked for every target.

# ── Age × Gender JOINT cells (8 cells) ────────────────────────────────────────
# Ensures each age group hits correct Male/Female percentage internally.
# 2024 Exit Poll / CPS gender distribution by age cohort:
#   18-29: 51% M / 49% F    30-44: 48% M / 52% F
#   45-64: 48% M / 52% F    65+:   47% M / 53% F
# Age proportions from 2024 CNN/Edison Exit Poll (matches Pollfish screenshots)
AGE_GENDER_TARGETS = {
    "18-29_Male":   0.0816,   # 16% × 0.51
    "18-29_Female": 0.0784,   # 16% × 0.49
    "30-44_Male":   0.1200,   # 25% × 0.48
    "30-44_Female": 0.1300,   # 25% × 0.52
    "45-64_Male":   0.1536,   # 32% × 0.48
    "45-64_Female": 0.1664,   # 32% × 0.52
    "65+_Male":     0.1269,   # 27% × 0.47
    "65+_Female":   0.1431,   # 27% × 0.53
}

# Age marginals (derived from cells above; used for convergence checks)
AGE_TARGETS = {"18-29": 0.16, "30-44": 0.25, "45-64": 0.32, "65+": 0.27}

# Gender marginals (derived from cells above)
GENDER_TARGETS = {"Male": 0.480, "Female": 0.520}

# ── Race × Education (White split into College / No College) ──────────────────
# 2024 Exit Poll universe:
#   White ~71% total — CPS splits ~53% No College / 47% College within White
#   Black 11%, Hispanic 11%, Asian/Other 7%
RACE_EDU_TARGETS = {
    "White No College": 0.3863,   # 0.71 × 0.53
    "White College":    0.3237,   # 0.71 × 0.47
    "Black":            0.110,
    "Hispanic":         0.110,
    "Asian / Other":    0.070,
}

# ── Education 4-way (Pollfish platform benchmark — matches likely-voter universe)
EDUCATION_TARGETS = {
    "High school or less":        0.15,
    "Some college/assoc. degree": 0.42,
    "College graduate":           0.24,
    "Postgraduate study":         0.19,
}

# ── Region (Census 2023 adult population — from usweightsmarch23.py)
REGION_TARGETS = {
    "New England":                 0.048,
    "Mid-Atlantic":                0.150,
    "Southeast Atlantic":          0.170,
    "Appalachia / South Interior": 0.081,
    "Great Lakes":                 0.172,
    "Lower Midwest / Plains":      0.061,
    "Southwest":                   0.117,
    "West / Mountain / Pacific":   0.201,
}


# ── Vote 2024 INCLUDING "Did not vote" (FEC certified + CPS 35% non-voter share)
VOTE_2024_TARGETS = {
    "Donald Trump":  0.3544,   # 49.91% × 0.65
    "Kamala Harris": 0.3446,   # 48.39% × 0.65
    "Third party":   0.0110,   # 1.70%  × 0.65
    "Did not vote":  0.2900,   # CPS 2024
}

# ── 2024 recall calibration (voters only — FEC certified)
RECALL2024_TARGETS_VOTERS = {
    "Donald Trump":  0.4991,
    "Kamala Harris": 0.4839,
    "Third party":   0.0170,
}
RECALL2024_NONVOTER_SHARE = 0.35  # CPS 2024 ASEC

# ── Vote history buckets (Gallup / CPS blend)
VOTE_HISTORY_TARGETS = {
    "Consistent voter": 0.48,
    "Occasional voter": 0.32,
    "New / non-voter":  0.20,
}

# ── Gender × Education (binary) joint cells ──────────────────────────────────
# Source: CPS 2023 + ACS 2022 adult population estimates
# Male:   ~48.2% of adult electorate; College ~45% of males
# Female: ~51.8% of adult electorate; College ~52% of females
GENDER_EDU_TARGETS = {
    "Male_No College":   0.265,   # 0.482 × 0.55
    "Male_College":      0.217,   # 0.482 × 0.45
    "Female_No College": 0.249,   # 0.518 × 0.48
    "Female_College":    0.269,   # 0.518 × 0.52
}

# ── Groyper identity crosstab ─────────────────────────────────────────────────
# Q10: "Do you identify as a Groyper, an individual who supports Nick Fuentes?"
# Responses are simply Yes / No — pass through directly, no remapping needed.
_GROYPER_BUCKET_MAP = {
    "Yes": "Yes",
    "No":  "No",
}

# ==============================================================================
# 4) VOTE HISTORY BUCKET (Q2 multi-select scoring)
# ==============================================================================

def make_vote_history_bucket(row, q2_cols: dict) -> str:
    """
    Classify respondent as Consistent / Occasional / New voter.
    q2_cols: dict mapping year string → column name, e.g. {"2024": "Q2_Voted_2024"}
    """
    never_col = "Q2_Never_Voted"
    if str(row.get(never_col, "0")).strip() in ("1","1.0","Selected","True","Yes"):
        return "New / non-voter"
    count = sum(
        1 for col in q2_cols.values()
        if str(row.get(col, "0")).strip() in ("1","1.0","Selected","True","Yes")
    )
    if count >= 3: return "Consistent voter"
    if count >= 1: return "Occasional voter"
    return "New / non-voter"

# (Propensity seed weighting removed — raking initialises from uniform weights)

# ==============================================================================
# 6) RAKING (Wisconsin methodology — verbose convergence output)
# ==============================================================================

def compute_deff(weights):
    w = np.array(weights, dtype=float); n = len(w); s = w.sum()
    if s <= 0 or n == 0: return np.nan
    return n * (w**2).sum() / (s**2)

def effective_n(weights):
    w = np.array(weights, dtype=float); s2 = (w**2).sum()
    return (w.sum()**2) / s2 if s2 > 0 else np.nan

def kish_deff(weights) -> float:
    """Kish (1965) design effect approximation: 1 + CV²(w)."""
    w = np.asarray(weights, dtype=float)
    if w.mean() == 0: return np.nan
    cv2 = w.var() / w.mean()**2
    return 1.0 + cv2

def print_deff_summary(label, weights):
    deff = compute_deff(weights); eff  = effective_n(weights); n = len(weights)
    flag = " ⚠️  DEFF > 1.5 — consider tightening cap" if deff > 1.5 else " ✅"
    print(f"  {label}: n={n}  DEFF={deff:.3f}  Eff.N={eff:.0f}{flag}")

def rake_weights(df_in, benchmarks, max_iter=60, weight_col="weight_rv",
                 init_weights=None, verbose=True):
    df2 = df_in.copy()
    if init_weights is not None:
        df2[weight_col] = init_weights.reindex(df2.index).fillna(1.0).clip(lower=0.01)
        m = df2[weight_col].mean()
        if m > 0: df2[weight_col] /= m
    else:
        df2[weight_col] = 1.0

    vars_list = [v for v in benchmarks if v in df2.columns]
    if verbose:
        print(f"\n{'='*60}")
        print(f"  RAKING  |  n={len(df2):,}  |  dims: {', '.join(vars_list)}")
        print(f"{'='*60}")
        print(f"  {'Iter':>4}  {'Max Deviation':>14}  Status")
        print(f"  {'-'*4}  {'-'*14}  {'-'*30}")

    prev_dev = None
    for i in range(1, max_iter + 1):
        for var, targets in benchmarks.items():
            if var not in df2.columns: continue
            cur = df2.groupby(var)[weight_col].sum(); tot = float(cur.sum())
            if tot <= 0: continue
            cur = cur / tot
            for cat, tgt in targets.items():
                if cat in cur and float(cur[cat]) > 0:
                    df2.loc[df2[var] == cat, weight_col] *= tgt / float(cur[cat])

        # Convergence check
        max_dev = 0.0
        for var, targets in benchmarks.items():
            if var not in df2.columns: continue
            cur = df2.groupby(var)[weight_col].sum()
            tot = float(cur.sum())
            if tot <= 0: continue
            cur = cur / tot
            for cat, tgt in targets.items():
                max_dev = max(max_dev, abs(float(cur.get(cat, 0.0)) - tgt))

        converged = max_dev < 0.0001
        improving = prev_dev is None or max_dev < prev_dev
        status = "✅ CONVERGED" if converged else ("improving" if improving else "plateaued")
        if verbose and (converged or i <= 10 or i % 10 == 0):
            print(f"  {i:>4}  {max_dev:>13.6f}%  {status}")
        prev_dev = max_dev
        if converged: break

    if verbose:
        # Post-raking marginals
        print(f"\n  {'Variable':<18}  {'Category':<32}  {'Target':>7}  {'Actual':>7}  {'Diff':>7}")
        print(f"  {'-'*18}  {'-'*32}  {'-'*7}  {'-'*7}  {'-'*7}")
        for var, targets in benchmarks.items():
            if var not in df2.columns: continue
            cur = df2.groupby(var)[weight_col].sum()
            tot = float(cur.sum())
            if tot <= 0: continue
            cur = cur / tot
            for cat, tgt in targets.items():
                obs  = float(cur.get(cat, 0.0))
                diff = obs - tgt
                flag = " ⚠️" if abs(diff) > 0.005 else ""
                print(f"  {var:<18}  {cat:<32}  {tgt*100:>6.2f}%  {obs*100:>6.2f}%  {diff*100:>+6.2f}%{flag}")
        print(f"{'='*60}\n")
    return df2

# ==============================================================================
# 7) 2024 RECALL CALIBRATION (Wisconsin two-stage methodology)
# ==============================================================================

def apply_recall2024_calibration(df_in, bucket_col="Vote2024_Bucket",
                                  base_wcol="weight_rv", out_wcol="weight_rv",
                                  voter_targets=None, nonvoter_share=None,
                                  trim_pctl=0.990, max_multiplier=2.0):
    if voter_targets is None: voter_targets = RECALL2024_TARGETS_VOTERS
    if nonvoter_share is None: nonvoter_share = RECALL2024_NONVOTER_SHARE
    df2 = df_in.copy()
    if base_wcol not in df2.columns or bucket_col not in df2.columns:
        return df2

    voter_keys   = list(voter_targets.keys())
    nonvoter_key = "Did not vote"

    # Stage 1: voters → FEC popular vote shares
    voter_mask = df2[bucket_col].isin(voter_keys)
    if voter_mask.sum() > 0:
        obs = df2.loc[voter_mask].groupby(bucket_col)[base_wcol].sum().reindex(voter_keys).fillna(0.0)
        obs_s = obs / float(obs.sum()) if obs.sum() > 0 else obs
        mult = {k: min(voter_targets[k] / max(float(obs_s.get(k, 1e-9)), 1e-9), max_multiplier)
                for k in voter_keys}
        df2[out_wcol] = df2[base_wcol].astype(float)
        df2.loc[voter_mask, out_wcol] = (
            df2.loc[voter_mask, base_wcol].astype(float)
            * df2.loc[voter_mask, bucket_col].map(mult).astype(float)
        )

    # Stage 2: non-voter share → CPS estimate
    nv_mask  = df2[bucket_col] == nonvoter_key
    all_mask = df2[bucket_col].isin(voter_keys + [nonvoter_key])
    if nv_mask.sum() > 0 and all_mask.sum() > 0:
        tot_w = float(df2.loc[all_mask, out_wcol].sum())
        nv_w  = float(df2.loc[nv_mask,  out_wcol].sum())
        cur_nv = nv_w / tot_w if tot_w > 0 else 0.0
        if cur_nv > 0:
            df2.loc[nv_mask, out_wcol] *= min(nonvoter_share / cur_nv, max_multiplier)

    # Renormalize
    base_s = float(df2[base_wcol].sum()); new_s = float(df2[out_wcol].sum())
    if new_s > 0 and base_s > 0: df2[out_wcol] *= base_s / new_s

    # Trim
    if trim_pctl is not None:
        hi = float(df2[out_wcol].quantile(trim_pctl))
        if hi > 0: df2[out_wcol] = df2[out_wcol].clip(upper=hi)
        new_s2 = float(df2[out_wcol].sum())
        if new_s2 > 0 and base_s > 0: df2[out_wcol] *= base_s / new_s2
    return df2

# ==============================================================================
# 8) LIKELY VOTER MODEL (arithmetic mean — Wisconsin Improvement #1)
# ==============================================================================

Q3_LV_WEIGHTS = {
    "I am certain to vote and highly motivated to do so":          1.0000,
    "I am very likely to vote and feel motivated":                  0.9000,
    "I am somewhat likely to vote but not strongly motivated":      0.2000,
    "I am motivated but unsure if I will actually vote":            0.1000,
    "I am not very likely to vote and feel little motivation":      0.0150,
    "I am certain not to vote":                                     0.0001,
}
Q4_LV_WEIGHTS = {
    "In person on Election Day — I know my polling location":                        1.0000,
    "Early in-person voting — I know when and where early voting is available":      1.0000,
    "Mail-in or absentee ballot — I have already requested or received my ballot":   1.0000,
    "In person on Election Day — I still need to confirm my polling location":       0.5000,
    "Early in-person voting — I still need to look up early voting details":         0.5000,
    "Mail-in or absentee ballot — I plan to request one but haven't yet":            0.5000,
    "I haven't decided how I will vote yet":                                         0.2000,
    "I do not plan to vote":                                                         0.0050,
}
Q5_LV_WEIGHTS = {
    "All or nearly all of them": 1.0000,
    "Most of them":              0.9500,
    "About half":                0.2000,
    "A few of them":             0.0150,
    "Not sure":                  0.0500,
    "None of them":              0.0001,
}

def compute_lv_score(row, q2_cols: dict) -> float:
    q3wt = Q3_LV_WEIGHTS.get(str(row.get("Q3_VoteIntent","")).strip(), 0.0)
    q4wt = Q4_LV_WEIGHTS.get(str(row.get("Q4_BallotMethod","")).strip(), 0.0)
    q5wt = Q5_LV_WEIGHTS.get(str(row.get("Q5_SocialVote","")).strip(), 0.0)
    # Q2 history score
    count = sum(
        1 for col in q2_cols.values()
        if str(row.get(col, "0")).strip() in ("1","1.0","Selected","True","Yes")
    )
    history_table = [0.10, 0.40, 0.60, 0.75, 0.85, 0.93, 1.00]
    history = history_table[min(count, 6)]
    propensity = 0.50 * q3wt + 0.30 * q4wt + 0.20 * q5wt
    return float(np.clip(0.60 * propensity + 0.40 * history, 0.0, 1.0))

# ==============================================================================
# 9) BOOTSTRAP SE (Wisconsin Improvement #6)
# ==============================================================================

def bootstrap_se(df_in, weight_col, question_col, response, n_boot=500, seed=42):
    rng = np.random.default_rng(seed)
    idx = np.arange(len(df_in))
    q_arr = df_in[question_col].astype(str).str.strip().to_numpy()
    w_arr = pd.to_numeric(df_in[weight_col], errors="coerce").fillna(0.0).to_numpy()
    estimates = []
    for _ in range(n_boot):
        bi = rng.choice(idx, size=len(idx), replace=True)
        bw = w_arr[bi]; tot = bw.sum()
        if tot > 0: estimates.append(bw[q_arr[bi] == response].sum() / tot * 100.0)
    return float(np.std(estimates)) if estimates else np.nan

# ==============================================================================
# METHOD 5 — ENTROPY BALANCING
# Hainmueller (2012) — guarantees exact moment matching while minimising
# KL-divergence from uniform weights. Falls back to raking if it fails.
# ==============================================================================

def entropy_balance(df_in, targets, weight_col="weight_rv", max_iter=500, tol=1e-6, verbose=True):
    df2 = df_in.copy()
    n = len(df2)
    cols, t_vec = [], []
    for var, cat_tgts in targets.items():
        if var not in df2.columns: continue
        for cat, tgt in cat_tgts.items():
            indicator = (df2[var].astype(str).str.strip() == str(cat)).astype(float).values
            if indicator.sum() < 1: continue
            cols.append(indicator)
            t_vec.append(tgt)
    if not cols:
        df2[weight_col] = 1.0
        return df2
    X = np.column_stack(cols)
    t = np.array(t_vec)

    def objective(lam):
        scores = X @ lam
        scores_c = scores - scores.max()
        log_Z  = scores.max() + np.log(np.exp(scores_c).sum())
        return -(log_Z - lam @ t * n)

    def gradient(lam):
        scores = X @ lam
        scores_c = scores - scores.max()
        exp_s  = np.exp(scores_c)
        w_norm = exp_s / exp_s.sum()
        return -(n * (X.T @ w_norm) - t * n)

    try:
        result = minimize(objective, np.zeros(len(t)), jac=gradient, method="L-BFGS-B",
                          options={"maxiter": max_iter, "ftol": tol, "gtol": 1e-7})
        if not result.success and result.fun > 1.0:
            raise ValueError(f"EB did not converge: {result.message}")
        _s = X @ result.x
        raw_w = np.exp(_s - _s.max())
        raw_w = np.clip(raw_w, a_min=None, a_max=3.5 * raw_w.mean())
        df2[weight_col] = raw_w / raw_w.mean()
    except Exception as e:
        if verbose:
            print(f"  ⚠️  Entropy balancing failed ({e}), falling back to raking")
        df2 = rake_weights(df_in, targets, weight_col=weight_col, verbose=verbose)
    return df2

# ==============================================================================
# METHOD 8 — RESPONSE PROPENSITY SCORE
# ==============================================================================

def compute_propensity_scores(df_in, targets, prop_vars):
    cols_available = [v for v in prop_vars if v in df_in.columns]
    if not cols_available: return pd.Series(np.nan, index=df_in.index)
    X_parts = []
    for var in cols_available:
        dummies = pd.get_dummies(df_in[var].astype(str).str.strip(), prefix=var, drop_first=False)
        X_parts.append(dummies)
    X = pd.concat(X_parts, axis=1).fillna(0).astype(float)
    pop_rows = []
    n_pop = min(len(df_in) * 5, 5000)
    for var in cols_available:
        if var not in targets: continue
        for cat, prop in targets[var].items():
            count = int(np.round(n_pop * prop / len(targets)))
            for _ in range(max(1, count)):
                row = {f"{v}_{c}": 0 for v in cols_available
                       for c in df_in[v].astype(str).str.strip().unique()}
                key = f"{var}_{cat}"
                if key in X.columns: row[key] = 1
                pop_rows.append(row)
    pop_df = pd.DataFrame(pop_rows, columns=X.columns).fillna(0)
    y = np.concatenate([np.ones(len(X)), np.zeros(len(pop_df))])
    X_all = pd.concat([X, pop_df], ignore_index=True).fillna(0)
    try:
        clf = LogisticRegression(max_iter=300, C=1.0, solver="lbfgs", class_weight="balanced")
        clf.fit(X_all, y)
        return pd.Series(clf.predict_proba(X)[:, 1], index=df_in.index)
    except Exception:
        return pd.Series(np.nan, index=df_in.index)

# ==============================================================================
# METHOD 10 — INTRACLASS CORRELATION (ICC) / GEOGRAPHIC CLUSTERING EFFECT
# ==============================================================================

def compute_icc(df_in, cluster_col, outcome_col, weight_col="weight_rv"):
    if cluster_col not in df_in.columns or outcome_col not in df_in.columns:
        return {"ICC": np.nan, "DEFF_cluster": np.nan, "mean_cluster_size": np.nan}
    df2 = df_in[[cluster_col, outcome_col, weight_col]].copy().dropna()
    uniq = [u for u in df2[outcome_col].astype(str).str.strip().unique()
            if u not in ("nan","","None")]
    if not uniq: return {"ICC": np.nan, "DEFF_cluster": np.nan, "mean_cluster_size": np.nan}
    y = (df2[outcome_col].astype(str).str.strip() == sorted(uniq)[0]).astype(float).values
    w = pd.to_numeric(df2[weight_col], errors="coerce").fillna(1.0).values
    clusters = df2[cluster_col].astype(str).str.strip().values
    grand_mean = np.average(y, weights=w)
    unique_clusters = np.unique(clusters)
    J = len(unique_clusters)
    if J < 2: return {"ICC": np.nan, "DEFF_cluster": np.nan, "mean_cluster_size": np.nan}
    SSB = SSW = n_total = 0.0
    cluster_sizes = []
    for cl in unique_clusters:
        mask = clusters == cl
        yj, wj = y[mask], w[mask]
        nj = wj.sum()
        if nj == 0: continue
        yj_bar = np.average(yj, weights=wj)
        SSB += nj * (yj_bar - grand_mean)**2
        SSW += np.sum(wj * (yj - yj_bar)**2)
        n_total += nj
        cluster_sizes.append(nj)
    MSB = SSB / (J - 1) if J > 1 else np.nan
    MSW = SSW / (n_total - J) if n_total > J else np.nan
    if not MSW or MSW <= 0 or np.isnan(MSW):
        return {"ICC": np.nan, "DEFF_cluster": np.nan,
                "mean_cluster_size": round(float(np.mean(cluster_sizes)), 1)}
    k_bar = float(np.mean(cluster_sizes))
    denom = MSB + (k_bar - 1) * MSW
    icc   = float(np.clip((MSB - MSW) / denom if denom > 0 else 0.0, 0, 1))
    return {"ICC": round(icc, 4), "DEFF_cluster": round(1 + (k_bar - 1) * icc, 3),
            "mean_cluster_size": round(k_bar, 1), "n_clusters": J}

# ==============================================================================
# METHOD 11 — TAYLOR SERIES LINEARISATION MARGIN OF ERROR
# ==============================================================================

def compute_moe(p, n, deff=1.0, conf=0.95):
    z = {0.90: 1.645, 0.95: 1.960, 0.99: 2.576}.get(conf, 1.960)
    if n <= 1 or deff <= 0 or not (0 < p < 1): return np.nan
    return z * np.sqrt(deff * p * (1 - p) / n) * 100

# ==============================================================================
# METHOD 12 — CELL COLLAPSE SAFEGUARD
# ==============================================================================

def collapse_thin_cells(df_in, targets, min_cell_n=5):
    df2 = df_in.copy()
    tgts2 = deepcopy(targets)
    for var, cat_tgts in list(tgts2.items()):
        if var not in df2.columns: continue
        counts = df2[var].astype(str).str.strip().value_counts()
        thin = [c for c, n in counts.items() if n < min_cell_n and c in cat_tgts]
        if not thin: continue
        for tc in thin:
            remaining = {c: t for c, t in cat_tgts.items()
                         if c != tc and c in counts and counts.get(c, 0) >= min_cell_n}
            if not remaining: continue
            tc_tgt = cat_tgts[tc]
            neighbour = min(remaining, key=lambda c: abs(remaining[c] - tc_tgt))
            df2.loc[df2[var].astype(str).str.strip() == tc, var] = neighbour
            tgts2[var][neighbour] = tgts2[var].get(neighbour, 0.0) + tc_tgt
            del tgts2[var][tc]
            print(f"  ⚠️  [{var}] cell '{tc}' (n={counts.get(tc, 0)}) collapsed into '{neighbour}'")
    return df2, tgts2

# ==============================================================================
# DIAGNOSTIC REPORTING (Methods 9 & 11)
# ==============================================================================

def weight_diagnostics_report(df_in, weight_col, targets, n_unweighted, propensity_col=None):
    w = pd.to_numeric(df_in[weight_col], errors="coerce").fillna(0).values
    d = {
        "n_unweighted":       n_unweighted,
        "n_weighted_approx":  int(w.sum()),
        "DEFF":               round(compute_deff(w), 3),
        "effective_N":        round(effective_n(w), 1),
        "Kish_DEFF":          round(kish_deff(w), 3),
        "weight_min":         round(float(w.min()), 3),
        "weight_max":         round(float(w.max()), 3),
        "weight_mean":        round(float(w.mean()), 3),
        "weight_median":      round(float(np.median(w)), 3),
        "weight_p99":         round(float(np.percentile(w, 99)), 3),
        "pct_weight_gt2":     round(float((w > 2.0).mean() * 100), 1),
        "pct_weight_gt3":     round(float((w > 3.0).mean() * 100), 1),
    }
    if propensity_col and propensity_col in df_in.columns:
        ps = pd.to_numeric(df_in[propensity_col], errors="coerce").dropna().values
        d["propensity_mean"] = round(float(ps.mean()), 3)
        d["pct_low_propensity"] = round(float((ps < 0.10).mean() * 100), 1)
        d["n_extreme_non_responders"] = int((ps < 0.05).sum())
    return d

def covariate_balance_table(df_in, targets, weight_col):
    rows = []
    for var, cat_tgts in targets.items():
        if var not in df_in.columns: continue
        w = pd.to_numeric(df_in[weight_col], errors="coerce").fillna(0.0)
        tot = w.sum()
        for cat, tgt in cat_tgts.items():
            obs_w   = float(w[df_in[var].astype(str).str.strip() == str(cat)].sum())
            obs_pct = obs_w / tot if tot > 0 else 0.0
            diff    = obs_pct - tgt
            smd     = abs(diff) / np.sqrt(tgt*(1-tgt)) if 0 < tgt < 1 else np.nan
            rows.append({
                "Variable":   var, "Category": cat,
                "Target %":   round(tgt*100, 2),
                "Weighted %": round(obs_pct*100, 2),
                "Diff (pp)":  round(diff*100, 2),
                "SMD":        round(smd, 3) if not np.isnan(smd) else np.nan,
                "Balanced?":  "✅" if (not np.isnan(smd) and smd < 0.10) else "⚠️",
            })
    return pd.DataFrame(rows)

# ==============================================================================
# 10) DESIGN SYSTEM (identical to Wisconsin/US weights)
# ==============================================================================

_C_TITLE_BG="0D1B2A"; _C_TITLE_FG="FFFFFF"
_C_GRP_BG="1B3A5C";   _C_GRP_FG="FFFFFF"
_C_SUB_BG="274D6E";   _C_SUB_FG="E8F0F7"
_C_TOTAL_BG="1B3A5C"; _C_TOTAL_FG="FFFFFF"
_C_ROW_ODD="F7F9FC";  _C_ROW_EVEN="FFFFFF"
_C_LABEL_FG="0D1B2A"; _C_DATA_FG="1B2A3B"
_C_BORDER="C5D3E0";   _C_TOTAL_COL="EFF4F9"
_C_LOW_N_BG="FFE0B2"; _C_LOW_N_FG="7B3F00"
LOW_N_THRESHOLD = 30

_HEAT = [
    (80,"1A4F8A","FFFFFF"),(60,"2E6EB0","FFFFFF"),(40,"5B9BD5","FFFFFF"),
    (25,"9DC3E6","1B2A3B"),(10,"D6E8F5","1B2A3B"),( 0,"F7F9FC","1B2A3B"),
]
_NET_HEAT_POS = [
    (60,"1A6B3C","FFFFFF"),(40,"2E8B57","FFFFFF"),(20,"52A97A","FFFFFF"),
    ( 5,"90CBB0","1B2A3B"),( 0,"C8E6D4","1B2A3B"),
]
_NET_HEAT_NEG = [
    (-60,"8B1A1A","FFFFFF"),(-40,"B03030","FFFFFF"),(-20,"CC5555","FFFFFF"),
    ( -5,"E6AAAA","1B2A3B"),(  0,"F7DCDC","1B2A3B"),
]
_GROUP_ACCENTS = [
    "1B3A5C","204060","163352","1A3D68","183558",
    "1C3F6A","17325A","1B3A5C","204060","163352","1A3D68",
]

def _heat_fill(v):
    try: v = float(v)
    except: return (_C_ROW_ODD, _C_DATA_FG)
    for thr,bg,fg in _HEAT:
        if v >= thr: return (bg,fg)
    return (_C_ROW_ODD, _C_DATA_FG)

def _net_heat_fill(v):
    try: v = float(v)
    except: return (_C_ROW_ODD, _C_DATA_FG)
    if v >= 0:
        for thr,bg,fg in _NET_HEAT_POS:
            if v >= thr: return (bg,fg)
    else:
        for thr,bg,fg in _NET_HEAT_NEG:
            if v <= thr: return (bg,fg)
    return (_C_ROW_ODD, _C_DATA_FG)

# Party heat: RED for Republican lead (positive), BLUE for Democrat lead (negative)
_PARTY_HEAT_REP = [
    (60, "8B1A1A", "FFFFFF"), (40, "B03030", "FFFFFF"),
    (20, "CC5555", "FFFFFF"), (5,  "E6AAAA", "1B2A3B"),
    (0,  "F7DCDC", "1B2A3B"),
]
_PARTY_HEAT_DEM = [
    (60, "0D2D6B", "FFFFFF"), (40, "1A4F8A", "FFFFFF"),
    (20, "2E6EB0", "FFFFFF"), (5,  "9DC3E6", "1B2A3B"),
    (0,  "D6E8F5", "1B2A3B"),
]

def _party_heat_fill(v):
    try: v = float(v)
    except: return (_C_ROW_ODD, _C_DATA_FG)
    if v >= 0:
        for thr,bg,fg in _PARTY_HEAT_REP:
            if v >= thr: return (bg,fg)
        return (_PARTY_HEAT_REP[-1][1], _PARTY_HEAT_REP[-1][2])
    else:
        abs_v = abs(v)
        for thr,bg,fg in _PARTY_HEAT_DEM:
            if abs_v >= thr: return (bg,fg)
        return (_PARTY_HEAT_DEM[-1][1], _PARTY_HEAT_DEM[-1][2])

def _party_net_label(val):
    if not isinstance(val, (int, float, np.floating)) or np.isnan(val): return ""
    abs_val = abs(val)
    if val > 0.05: return f"R+{abs_val:.1f}"
    elif val < -0.05: return f"D+{abs_val:.1f}"
    return "EVEN"

_pf = lambda h: PatternFill("solid", fgColor=h)
def _bd(c=None):
    c = c or _C_BORDER; s = Side(style="thin", color=c)
    return Border(left=s,right=s,top=s,bottom=s)
def _left_bd():
    thick=Side(style="medium",color="8BAFCC"); thin=Side(style="thin",color=_C_BORDER)
    return Border(left=thick,right=thin,top=thin,bottom=thin)

_AC = Alignment(horizontal="center",vertical="center",wrap_text=True)
_AL = Alignment(horizontal="left",  vertical="center",wrap_text=False)
_H_TITLE=22; _H_GROUP=18; _H_SUBHDR=30; _H_DATA=16; _H_N_ROW=12; _H_NET_ROW=18

# ── Net row helpers (Wisconsin) ───────────────────────────────────────────────

def _compute_net_values(block_df, approve_keys, disapprove_keys):
    results = {}
    value_cols = [c for c in block_df.columns if c != "Response"]
    for col in value_cols:
        app = sum(float(block_df.loc[block_df["Response"] == k, col].sum())
                  for k in approve_keys    if k in block_df["Response"].values)
        dis = sum(float(block_df.loc[block_df["Response"] == k, col].sum())
                  for k in disapprove_keys if k in block_df["Response"].values)
        results[col] = (app, dis, app - dis)
    return results

def _write_net_rows(ws, r, block_df, col_groups,
                    approve_keys, disapprove_keys,
                    col_idx_to_var_cat, group_starts):
    ncols = block_df.shape[1]
    nets  = _compute_net_values(block_df, approve_keys, disapprove_keys)
    ROW_DEFS = [
        ("Approve",    0, "1A6B3C", "FFFFFF", False),
        ("Disapprove", 1, "8B1A1A", "FFFFFF", False),
        ("Net (App – Disapp)", 2, "1A2A3A", "FFFFFF", True),
    ]
    for label, val_idx, lbl_bg, lbl_fg, use_net_heat in ROW_DEFS:
        ws.row_dimensions[r].height = _H_NET_ROW
        lc = ws.cell(r, 1, value=label)
        lc.fill = _pf(lbl_bg)
        lc.font = Font(color=lbl_fg, bold=True, size=9, name="Calibri", italic=True)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        lc.border = _bd()
        for j in range(2, ncols + 1):
            col_name = block_df.columns[j - 1]
            val = nets[col_name][val_idx] if col_name in nets else ""
            cell = ws.cell(r, j, value=val if val != "" else "")
            cell.border = _left_bd() if j in group_starts else _bd()
            cell.alignment = _AC
            if isinstance(val, (int, float, np.floating)) and val == val:
                cell.number_format = '+0.0"%";-0.0"%";0.0"%"' if use_net_heat else '0.0"%"'
                if use_net_heat:
                    bg_h, fg_h = _net_heat_fill(val)
                else:
                    bg_h, fg_h = (_net_heat_fill(val) if val_idx == 0 else _net_heat_fill(-val))
                cell.fill = _pf(bg_h)
                cell.font = Font(color=fg_h, bold=use_net_heat, size=9, name="Calibri", italic=True)
            else:
                cell.fill = _pf("E8EFF6")
                cell.font = Font(color="888888", size=9, name="Calibri", italic=True)
        r += 1
    return r

def _write_generic_net_rows(ws, r, block_df, group_starts, row_defs):
    ncols = block_df.shape[1]
    for row_def in row_defs:
        label, values_dict, lbl_bg, lbl_fg, use_signed_fmt, use_net_heat, invert_heat = row_def[:7]
        party_net_fmt  = row_def[7] if len(row_def) > 7 else False
        party_pct_sign = row_def[8] if len(row_def) > 8 else None
        ws.row_dimensions[r].height = _H_NET_ROW
        lc = ws.cell(r, 1, value=label)
        lc.fill = _pf(lbl_bg)
        lc.font = Font(color=lbl_fg, bold=True, size=9, name="Calibri", italic=True)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        lc.border = _bd()
        for j in range(2, ncols + 1):
            col_name = block_df.columns[j - 1]
            val = values_dict.get(col_name, "")
            cell = ws.cell(r, j)
            cell.border = _left_bd() if j in group_starts else _bd()
            cell.alignment = _AC
            if party_net_fmt and isinstance(val, (int, float, np.floating)) and not np.isnan(val):
                label_str = _party_net_label(val)
                cell.value = label_str
                cell.number_format = "@"
                bg_h, fg_h = _party_heat_fill(val)
                cell.fill = _pf(bg_h)
                cell.font = Font(color=fg_h, bold=True, size=9, name="Calibri", italic=True)
            elif party_pct_sign is not None and isinstance(val, (int, float, np.floating)) and not np.isnan(val):
                heat_input = val * party_pct_sign
                cell.value = val
                cell.number_format = '0.0"%"'
                bg_h, fg_h = _party_heat_fill(heat_input)
                cell.fill = _pf(bg_h)
                cell.font = Font(color=fg_h, bold=False, size=9, name="Calibri", italic=True)
            elif isinstance(val, (int, float, np.floating)) and not np.isnan(val):
                cell.value = val
                cell.number_format = ('+0.0"%";-0.0"%";0.0"%"' if use_signed_fmt else '0.0"%"')
                heat_val = -val if invert_heat else val
                bg_h, fg_h = _net_heat_fill(heat_val) if use_net_heat else _heat_fill(val)
                cell.fill = _pf(bg_h)
                cell.font = Font(color=fg_h, bold=use_signed_fmt, size=9, name="Calibri", italic=True)
            else:
                cell.value = ""
                cell.fill = _pf("E8EFF6")
                cell.font = Font(color="888888", size=9, name="Calibri", italic=True)
        r += 1
    return r

def _write_horserace_net_rows(ws, r, block_df, group_starts, rep_keys, dem_keys):
    value_cols = [c for c in block_df.columns if c != "Response"]
    def _sum_keys(keys):
        return {col: sum(float(block_df.loc[block_df["Response"] == k, col].sum())
                         for k in keys if k in block_df["Response"].values)
                for col in value_cols}
    rep_vals = _sum_keys(rep_keys)
    dem_vals = _sum_keys(dem_keys)
    net_vals = {col: rep_vals[col] - dem_vals[col] for col in value_cols}
    row_defs = [
        ("Republican Total", rep_vals, "8B1A1A", "FFFFFF", False, False, False, False, +1),
        ("Democrat Total",   dem_vals, "1A4F8A", "FFFFFF", False, False, False, False, -1),
        ("Net (R+ / D+)",    net_vals, "1A2A3A", "FFFFFF", False, False, False, True,  None),
    ]
    return _write_generic_net_rows(ws, r, block_df, group_starts, row_defs)

def _write_morelikely_net_rows(ws, r, block_df, group_starts, more_keys, less_keys):
    value_cols = [c for c in block_df.columns if c != "Response"]
    def _sum_keys(keys):
        return {col: sum(float(block_df.loc[block_df["Response"] == k, col].sum())
                         for k in keys if k in block_df["Response"].values)
                for col in value_cols}
    more_vals = _sum_keys(more_keys)
    less_vals = _sum_keys(less_keys)
    net_vals  = {col: more_vals[col] - less_vals[col] for col in value_cols}
    row_defs = [
        ("More Likely",      more_vals, "1A6B3C", "FFFFFF", False, False, False),
        ("Less Likely",      less_vals, "8B1A1A", "FFFFFF", False, False, True),
        ("Net (More – Less)", net_vals, "1A2A3A", "FFFFFF", True,  True,  False),
    ]
    return _write_generic_net_rows(ws, r, block_df, group_starts, row_defs)

def detect_horserace_keys(options):
    rep_keys, dem_keys = [], []
    for opt in options:
        lo = opt.lower()
        if any(t in lo for t in ["republican", " (r)", "(r) "]) or opt.strip().endswith("(R)"):
            rep_keys.append(opt)
        elif any(t in lo for t in ["democrat", "democratic", " (d)", "(d) "]) or opt.strip().endswith("(D)"):
            dem_keys.append(opt)
    return rep_keys, dem_keys

# ── Diagnostic Excel sheets ────────────────────────────────────────────────────

_C_HDR_BG  = "0D1B2A"; _C_HDR_FG  = "FFFFFF"
_C_SUB_BG2 = "274D6E"; _C_SUB_FG2 = "E8F0F7"
_C_ODD2    = "F7F9FC"; _C_EVEN2   = "FFFFFF"
_C_WARN_BG = "FFE0B2"; _C_WARN_FG = "7B3F00"
_C_OK_BG   = "C8E6D4"; _C_OK_FG   = "1A4F2A"

def _write_diag_header(ws, row, title, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, value=title)
    c.fill = _pf(_C_HDR_BG)
    c.font = Font(color=_C_HDR_FG, bold=True, size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1

def write_diagnostics_sheet(writer, diag_rv, diag_lv, balance_rv, balance_lv,
                              icc_results, moe_results):
    ws = writer.book.create_sheet("Weight Diagnostics")
    r  = 1
    r = _write_diag_header(ws, r, "Weighting Diagnostics — Summary Statistics", 4)
    for j, h in enumerate(["Statistic", "RV Value", "LV Value", "Notes"], 1):
        c = ws.cell(r, j, value=h)
        c.fill = _pf(_C_SUB_BG2); c.font = Font(color=_C_SUB_FG2, bold=True, size=9, name="Calibri")
        c.alignment = _AC; c.border = _bd()
    r += 1
    stat_notes = {
        "n_unweighted":             "Raw respondents used",
        "DEFF":                     "Design effect: 1.0=ideal, >2.0=concerning",
        "effective_N":              "n / DEFF — real statistical power",
        "Kish_DEFF":                "Kish 1+CV²(w) approximation",
        "weight_max":               "Largest individual weight",
        "pct_weight_gt2":           "% respondents with weight >2.0",
        "propensity_mean":          "Avg propensity score (0.5=balanced)",
        "pct_low_propensity":       "% respondents with propensity <0.10",
        "n_extreme_non_responders": "Respondents with propensity <0.05",
    }
    for i, k in enumerate(sorted(set(list(diag_rv.keys()) + list(diag_lv.keys())))):
        ws.row_dimensions[r].height = 16
        row_bg = _C_ODD2 if i % 2 == 0 else _C_EVEN2
        for j, val in enumerate([k, diag_rv.get(k,"—"), diag_lv.get(k,"—"), stat_notes.get(k,"")], 1):
            c = ws.cell(r, j, value=val); c.fill = _pf(row_bg)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center") if j in (1,4) else _AC
            c.border = _bd()
        r += 1
    r += 2
    r = _write_diag_header(ws, r, "Geographic Clustering (ICC) Analysis", 4)
    for j, h in enumerate(["Metric", "Value", "", "Notes"], 1):
        c = ws.cell(r, j, value=h); c.fill = _pf(_C_SUB_BG2)
        c.font = Font(color=_C_SUB_FG2, bold=True, size=9, name="Calibri")
        c.alignment = _AC; c.border = _bd()
    r += 1
    icc_notes = {
        "ICC":               "Intraclass correlation; 0=no clustering, 1=perfect clustering",
        "DEFF_cluster":      "DEFF from clustering alone: 1 + (m̅-1)*ICC",
        "mean_cluster_size": "Average respondents per region/cluster",
        "n_clusters":        "Number of geographic clusters observed",
    }
    for i, (k, v) in enumerate(icc_results.items()):
        ws.row_dimensions[r].height = 16
        row_bg = _C_ODD2 if i % 2 == 0 else _C_EVEN2
        for j, val in enumerate([k, v, "", icc_notes.get(k,"")], 1):
            c = ws.cell(r, j, value=val); c.fill = _pf(row_bg)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center") if j in (1,4) else _AC
            c.border = _bd()
        r += 1
    r += 2
    r = _write_diag_header(ws, r, "Covariate Balance After Weighting (SMD < 0.10 = balanced)", 7)
    for section, bal_df in [("RV", balance_rv), ("LV", balance_lv)]:
        ws.cell(r, 1, value=f"── {section} ──").font = Font(bold=True, color="163352", size=10, name="Calibri")
        r += 1
        for j, h in enumerate(bal_df.columns, 1):
            c = ws.cell(r, j, value=h); c.fill = _pf(_C_SUB_BG2)
            c.font = Font(color=_C_SUB_FG2, bold=True, size=9, name="Calibri")
            c.alignment = _AC; c.border = _bd()
        r += 1
        for i, row_data in bal_df.iterrows():
            ws.row_dimensions[r].height = 15
            row_bg = _C_ODD2 if i % 2 == 0 else _C_EVEN2
            for j, val in enumerate(row_data.values, 1):
                c = ws.cell(r, j, value=val); c.fill = _pf(row_bg)
                c.font = Font(size=9, name="Calibri"); c.alignment = _AC; c.border = _bd()
                if j == 7:
                    if val == "✅": c.fill = _pf(_C_OK_BG); c.font = Font(color=_C_OK_FG, size=9, bold=True, name="Calibri")
                    else: c.fill = _pf(_C_WARN_BG); c.font = Font(color=_C_WARN_FG, size=9, bold=True, name="Calibri")
            r += 1
        r += 2
    if moe_results is not None and len(moe_results) > 0:
        r = _write_diag_header(ws, r, "Margin of Error by Question (Taylor Series, DEFF-adjusted, 95% CI)", 5)
        for j, h in enumerate(moe_results.columns, 1):
            c = ws.cell(r, j, value=h); c.fill = _pf(_C_SUB_BG2)
            c.font = Font(color=_C_SUB_FG2, bold=True, size=9, name="Calibri")
            c.alignment = _AC; c.border = _bd()
        r += 1
        for i, row_data in moe_results.iterrows():
            ws.row_dimensions[r].height = 15
            row_bg = _C_ODD2 if i % 2 == 0 else _C_EVEN2
            for j, val in enumerate(row_data.values, 1):
                c = ws.cell(r, j, value=val); c.fill = _pf(row_bg)
                c.font = Font(size=9, name="Calibri"); c.alignment = _AC; c.border = _bd()
            r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50
    ws.sheet_properties.tabColor = "0D1B2A"

def write_bootstrap_sheet(writer, boot_results, rv_df, weight_col, q_meta):
    ws = writer.book.create_sheet("Bootstrap CI")
    r  = 1
    r  = _write_diag_header(ws, r, "Bootstrap Uncertainty Quantification (95% credible intervals)", 7)
    for j, h in enumerate(["Question","Response","Weighted %","Boot Mean","CI Low (2.5%)","CI High (97.5%)","SE"], 1):
        c = ws.cell(r, j, value=h); c.fill = _pf(_C_SUB_BG2)
        c.font = Font(color=_C_SUB_FG2, bold=True, size=9, name="Calibri")
        c.alignment = _AC; c.border = _bd()
    r += 1
    for q, resp_data in boot_results.items():
        if q in rv_df.columns:
            s = rv_df[q].astype(str).str.strip()
            w = pd.to_numeric(rv_df[weight_col], errors="coerce").fillna(0)
            wn = {}
            for rsp, wv in zip(s, w):
                if rsp not in ("nan","","None"): wn[rsp] = wn.get(rsp, 0.0) + float(wv)
            tot    = sum(wn.values())
            actual = {rsp: wv/tot*100 for rsp, wv in wn.items()} if tot > 0 else {}
        else:
            actual = {}
        title = q_meta.get(q, q)
        for i, (resp, stats) in enumerate(sorted(resp_data.items(), key=lambda x: -x[1]["mean"])):
            ws.row_dimensions[r].height = 15
            row_bg = _C_ODD2 if i % 2 == 0 else _C_EVEN2
            vals = [title if i == 0 else "", resp,
                    round(actual.get(resp, np.nan), 1),
                    round(stats["mean"], 1), round(stats["lo"], 1),
                    round(stats["hi"], 1),  round(stats["se"], 2)]
            for j, val in enumerate(vals, 1):
                c = ws.cell(r, j, value=val); c.fill = _pf(row_bg)
                c.font = Font(size=9, name="Calibri"); c.alignment = _AC; c.border = _bd()
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 48
    for ci in range(3, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.sheet_properties.tabColor = "0D1B2A"

# ==============================================================================
# 11) SURVEY META — question titles & response orders for THIS poll
# ==============================================================================

SURVEY_NAME   = "National Benchmark Survey — May 2025"
OUTPUT_FILE   = "weighted_outputs_national_may.xlsx"
TOPLINES_FILE = "toplines_national_may.txt"

APPROVE_RESPONSE_ORDER = [
    "Strongly approve","Somewhat approve",
    "Somewhat disapprove","Strongly disapprove","Neutral / no opinion",
]

QUESTION_ORDER_SINGLE = {
    "Q3_VoteIntent": {
        "title": "Q3 How would you describe your intention and motivation to vote in the 2026 Midterm Election?",
        "options": [
            "I am certain to vote and highly motivated to do so",
            "I am very likely to vote and feel motivated",
            "I am somewhat likely to vote but not strongly motivated",
            "I am motivated but unsure if I will actually vote",
            "I am not very likely to vote and feel little motivation",
            "I am certain not to vote",
        ],
    },
    "Q4_BallotMethod": {
        "title": "Q4 How do you plan to cast your ballot in the 2026 Midterm Election?",
        "options": [
            "In person on Election Day — I know my polling location",
            "In person on Election Day — I still need to confirm my polling location",
            "Early in-person voting — I know when and where early voting is available",
            "Early in-person voting — I still need to look up early voting details",
            "Mail-in or absentee ballot — I have already requested or received my ballot",
            "Mail-in or absentee ballot — I plan to request one but haven't yet",
            "I haven't decided how I will vote yet",
            "I do not plan to vote",
        ],
    },
    "Q5_SocialVote": {
        "title": "Q5 Thinking about the 5–10 people you are closest to, how many do you expect to vote?",
        "options": ["All or nearly all of them","Most of them","About half",
                    "A few of them","Not sure","None of them"],
    },
    "Q6_2024Vote": {
        "title": "Q6 Who did you vote for in the 2024 Presidential Election?",
        "options": ["Donald Trump","Kamala Harris","Third party","Did not vote"],
    },
    "Q8_PoliticalOutlook": {
        "title": "Q8 Please select the option that best describes your general political outlook.",
        "options": [
            "America First Republican (Borders, national sovereignty, trade, anti-establishment)",
            "Populist / Working-Class Republican (Jobs, wages, Social Security, cultural conservatism)",
            "Suburban / Professional Republican (Taxes, economic stability, socially moderate)",
            "Libertarian-Oriented Republican (Small government, low taxes, personal freedom)",
            "Progressive / Socialist Democrat (Healthcare, climate, inequality, student debt)",
            "Mainline / Institutional Democrat (Mainstream Democratic policies, steady governance)",
            "Working-Class / Union Democrat (Unions, wages, worker protections)",
            "Coalition / Civil Rights Democrat (Civil rights, education, community investment)",
            "Lean Republican Independent (Independent, usually votes Republican)",
            "Lean Democratic Independent (Independent, usually votes Democratic)",
            "Moderate Independent / Centrist (Issue-by-issue, ticket-splitting voter)",
            "Anti-Establishment Independent (Distrusts both parties, outsider-focused)",
            "None of these / No clear political preference",
        ],
    },
    "Q9_PartyID": {
        "title": "Q9 Which political party do you identify with?",
        "options": ["Democrat","Republican","Independent / Other"],
    },
    "Q10_Groyper": {
        "title": "Q10 Do you identify as a Groyper, an individual who supports Nick Fuentes?",
        "options": ["Yes", "No"],
    },
    "Q11_RightTrack": {
        "title": "Q11 Is the direction of the country on the right track or the wrong track?",
        "options": ["Right track","Wrong track","Not sure / no opinion"],
    },
    "Q13_GenericBallot": {
        "title": "Q13 If the 2026 Midterm elections were held today, who would you vote for?",
        "options": ["The Republican candidate","The Democrat candidate",
                    "A third-party / independent candidate","Undecided / Not sure"],
    },
    "Q14_TrumpApprove": {
        "title": "Q14 Do you approve or disapprove of Donald J. Trump's performance as President?",
        "options": APPROVE_RESPONSE_ORDER,
    },
    "Q18_TrumpConservatism": {
        "title": "Q18 How would you describe President Trump's actions and policies in office so far?",
        "options": ["Far too conservative","Somewhat too conservative","About the right balance",
                    "Somewhat too liberal","Far too liberal","Not sure / no opinion"],
    },
    "Q19_ForeignPolicy": {
        "title": "Q19 In your opinion, whose interests do President Trump's foreign policy decisions primarily serve?",
        "options": [
            "The American people above all else",
            "Mostly the American people, but with significant consideration for allies like Israel",
            "A balance between American interests and the interests of foreign allies",
            "Mostly the interests of foreign allies like Israel over the American people",
            "Foreign allies like Israel above the American people",
            "Not sure / no opinion",
        ],
    },
    "Q20_HHExpenses": {
        "title": "Q20 Over the last month, how difficult has it been to pay for your usual household expenses?",
        "options": ["Very difficult","Somewhat difficult","Not very difficult",
                    "Not at all difficult","Not sure / no opinion"],
    },
    "Q21_MassDeportation": {
        "title": "Q21 Do you support or oppose the mass deportation of all illegal immigrants?",
        "options": ["Strongly support","Somewhat support","Somewhat oppose",
                    "Strongly oppose","Not sure / no opinion"],
    },
    "Q22_GenPref": {
        "title": "Q22 When voting for a candidate, which generation do you most prefer?",
        "options": ["Gen Z (ages 18–28)","Millennial (ages 29–44)",
                    "Gen X (ages 45–60)","Baby Boomer (ages 61–79)","Not sure / no opinion"],
    },
    "Q23_PartnerQuality": {
        "title": "Q23 Which is the single most important quality when considering a long-term partner?",
        "options": ["Financial stability","Physical attractiveness","Emotional availability",
                    "Ambition / work ethic","Sense of humor","Little to no sexual experience",
                    "A lot of sexual experience","Religious affiliation","Not sure / no opinion"],
    },
    "Q24_IsraelPAC": {
        "title": "Q24 If a candidate accepted donations from a PAC that supports Israel, would that affect your vote?",
        "options": ["Much more likely","Somewhat more likely","Somewhat less likely",
                    "Much less likely","No difference / no opinion"],
    },
    "Q25_CharlieKirk": {
        "title": "Q25 Who do you believe was ultimately responsible for the assassination of Charlie Kirk?",
        "options": [
            "Tyler Robinson, acting alone",
            "Tyler Robinson, acting as a part of a larger organization",
            "A left-wing or anti-conservative organization",
            "A right-wing or anti-liberal organization",
            "Someone at Turning Point USA (TPUSA)",
            "A government or political actor",
            "Not sure / no opinion",
        ],
    },
}

PARTY_DETAILED_ORDER = QUESTION_ORDER_SINGLE["Q8_PoliticalOutlook"]["options"]
# Use hand-crafted short labels so every faction is clearly identifiable in column headers
PARTY_DETAILED_ORDER_DISPLAY = [
    "America First R",
    "Populist / WC R",
    "Suburban / Prof R",
    "Libertarian R",
    "Progressive D",
    "Mainline D",
    "Working-Class / Union D",
    "Coalition / Civil Rights D",
    "Lean R Independent",
    "Lean D Independent",
    "Moderate Independent",
    "Anti-Estab. Indep.",
    "No Clear Preference",
]
# Ensure lengths match (guard against future option list edits)
assert len(PARTY_DETAILED_ORDER_DISPLAY) == len(PARTY_DETAILED_ORDER), \
    "PARTY_DETAILED_ORDER_DISPLAY length must match PARTY_DETAILED_ORDER"
_party_display_map = dict(zip(PARTY_DETAILED_ORDER, PARTY_DETAILED_ORDER_DISPLAY))

Q2_ITEMS_DISPLAY = [
    ("Q2_Voted_2024", "2024 - Presidential Election"),
    ("Q2_Voted_2022", "2022 - Midterm Elections"),
    ("Q2_Voted_2020", "2020 - Presidential Election"),
    ("Q2_Voted_2018", "2018 - Midterm Elections"),
    ("Q2_Voted_2016", "2016 - Presidential Election"),
    ("Q2_Voted_2014", "2014 - Midterm Elections"),
    ("Q2_Never_Voted","I have never voted / was not eligible"),
]

# ==============================================================================
# 12) TABBOOK ENGINE (crosstab cache, block writer, matrix writer)
#     Identical design system to Wisconsin/US weights scripts
# ==============================================================================

_CURRENT_GROUP_ORDER_REF = []  # populated after GROUP_ORDER is defined

def build_unweighted_n_lookup(df_in, group_order):
    lookup = {}
    for _, var, cats in group_order:
        if var not in df_in.columns: continue
        s = df_in[var].astype(str).str.strip()
        for cat in cats:
            lookup[(var,cat)] = int((s == str(cat)).sum())
    return lookup

def _build_crosstab_cache(df_in, q_col, w_col, group_order):
    if q_col not in df_in.columns or len(df_in) == 0:
        return {}, {}
    w  = pd.to_numeric(df_in[w_col], errors="coerce").fillna(0.0).to_numpy()
    qv = df_in[q_col].astype(str).str.strip().replace({"nan":np.nan,"":np.nan}).to_numpy()
    tmp = pd.DataFrame({"_q":qv,"_w":w})
    tmp = tmp[tmp["_q"].notna()]
    t   = tmp.groupby("_q")["_w"].sum()
    tot = float(t.sum())
    overall = {r: float(t.get(r,0.0))/tot*100.0 for r in t.index} if tot > 0 else {}
    cell = {}
    seen = set()
    for _, var, _ in [(gl,v,_cat) for gl,v,cats in group_order for _cat in cats]:
        if var in seen or var not in df_in.columns: continue
        seen.add(var)
        vv = df_in[var].astype(str).str.strip().replace({"nan":np.nan,"":np.nan}).to_numpy()
        tmp2 = pd.DataFrame({"_var":vv,"_q":qv,"_w":w})
        tmp2 = tmp2[tmp2["_var"].notna() & tmp2["_q"].notna()]
        grp  = tmp2.groupby(["_var","_q"],observed=True)["_w"].sum()
        vtot = grp.groupby(level=0).sum()
        for (cv,rv),ws in grp.items():
            d = float(vtot.get(cv,0.0))
            cell[(var,cv,rv)] = (float(ws)/d*100.0) if d > 0 else 0.0
    return overall, cell

def build_block_table(df_in, q_col, w_col, group_order, ordered_responses):
    # Use composite key "var|||cat" as DataFrame column names to avoid collisions
    # when the same category label appears in multiple groups. Display label
    # is stripped at write time.
    flat = [(gl,v,c) for gl,v,cats in group_order for c in cats]
    overall, cell = _build_crosstab_cache(df_in, q_col, w_col, group_order)
    seen_keys: dict = {}
    unique_col_keys = []
    for _, var, cat in flat:
        k = f"{var}|||{cat}"
        if k in seen_keys:
            seen_keys[k] += 1
            unique_col_keys.append(f"{k}||#{seen_keys[k]}")
        else:
            seen_keys[k] = 0
            unique_col_keys.append(k)
    data: dict = {"Response": [], "Total": []}
    for ck in unique_col_keys:
        data[ck] = []
    for resp in ordered_responses:
        data["Response"].append(resp)
        data["Total"].append(float(overall.get(resp,0.0)))
        for ck, (_,var,cat) in zip(unique_col_keys, flat):
            data[ck].append(float(cell.get((var,cat,resp),0.0)))
    block_df = pd.DataFrame(data)
    col_groups = []
    cur = 3
    for gl,var,cats in group_order:
        col_groups.append((gl, cur, cur+len(cats)-1))
        cur += len(cats)
    return block_df, col_groups

def write_block_to_sheet(ws, start_row, title, block_df, col_groups, unweighted_n_lookup=None,
                          approve_keys=None, disapprove_keys=None,
                          rep_keys=None, dem_keys=None,
                          more_keys=None, less_keys=None):
    """Write a standard crosstab block with optional net rows."""
    ncols = block_df.shape[1]
    group_starts = {sc for _,sc,_ in col_groups}
    r = start_row

    col_idx_to_var_cat = {}
    ci = 3
    for _,var,cats in _CURRENT_GROUP_ORDER_REF:
        for cat in cats:
            col_idx_to_var_cat[ci] = (var,cat); ci += 1

    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)
    c = ws.cell(r,1,value=title)
    c.fill = _pf(_C_TITLE_BG); c.font = Font(color=_C_TITLE_FG,bold=True,size=11,name="Calibri")
    c.alignment = Alignment(horizontal="left",vertical="center",wrap_text=False,indent=1)
    ws.row_dimensions[r].height = _H_TITLE; r += 1

    ws.row_dimensions[r].height = _H_GROUP
    ws.cell(r,1).fill = _pf(_C_GRP_BG); ws.cell(r,1).border = _bd()
    tc = ws.cell(r,2,value="Total")
    tc.fill = _pf(_C_TOTAL_BG); tc.font = Font(color=_C_TOTAL_FG,bold=True,size=9,name="Calibri")
    tc.alignment = _AC; tc.border = _bd()
    for gi,(gl,sc,ec) in enumerate(col_groups):
        bg = _GROUP_ACCENTS[gi % len(_GROUP_ACCENTS)]
        if sc < ec: ws.merge_cells(start_row=r,start_column=sc,end_row=r,end_column=ec)
        cell = ws.cell(r,sc,value=str(gl))
        cell.fill = _pf(bg); cell.font = Font(color="FFFFFF",bold=True,size=9,name="Calibri")
        cell.alignment = _AC
        for cc in range(sc,ec+1):
            ws.cell(r,cc).fill = _pf(bg)
            ws.cell(r,cc).border = _left_bd() if cc == sc else _bd()
    r += 1

    # Sub-header — strip composite key prefix "var|||cat" → display just "cat"
    ws.row_dimensions[r].height = _H_SUBHDR
    for j,cn in enumerate(block_df.columns,start=1):
        display_cn = str(cn).split("|||")[-1].split("||#")[0] if "|||" in str(cn) else str(cn)
        cell = ws.cell(r,j,value=display_cn)
        cell.fill = _pf(_C_SUB_BG); cell.font = Font(color=_C_SUB_FG,bold=False,size=8,name="Calibri")
        cell.alignment = (Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1)
                          if j == 1 else _AC)
        cell.border = _left_bd() if j in group_starts else _bd()
    r += 1

    if unweighted_n_lookup is not None:
        ws.row_dimensions[r].height = _H_N_ROW
        nl = ws.cell(r,1,value="(unweighted n)")
        nl.fill = _pf("E8EFF6"); nl.font = Font(color="555555",size=7,italic=True,name="Calibri")
        nl.alignment = Alignment(horizontal="left",vertical="center",indent=1); nl.border = _bd()
        ws.cell(r,2,value="").border = _bd()
        for ji in range(3,ncols+1):
            vc = col_idx_to_var_cat.get(ji)
            if vc and vc in unweighted_n_lookup:
                nv = unweighted_n_lookup[vc]
                nc = ws.cell(r,ji,value=nv)
                if nv < LOW_N_THRESHOLD:
                    nc.fill = _pf(_C_LOW_N_BG); nc.font = Font(color=_C_LOW_N_FG,size=7,bold=True,name="Calibri")
                else:
                    nc.fill = _pf("F0F4F8"); nc.font = Font(color="555555",size=7,name="Calibri")
                nc.alignment = _AC; nc.border = _left_bd() if ji in group_starts else _bd()
            else:
                ws.cell(r,ji,value="").border = _bd()
        r += 1

    for i in range(block_df.shape[0]):
        ws.row_dimensions[r].height = _H_DATA
        row_bg = _C_ROW_ODD if i % 2 == 0 else _C_ROW_EVEN
        for j in range(1,ncols+1):
            val  = block_df.iloc[i,j-1]
            cell = ws.cell(r,j,value=val)
            cell.border = _left_bd() if j in group_starts else _bd()
            if j == 1:
                cell.fill = _pf(row_bg); cell.font = Font(color=_C_LABEL_FG,size=9,name="Calibri")
                cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1)
            elif j == 2:
                cell.fill = _pf(_C_TOTAL_COL); cell.font = Font(color=_C_DATA_FG,bold=True,size=9,name="Calibri")
                cell.alignment = _AC
                if isinstance(val,(int,float,np.floating)) and val == val:
                    cell.number_format = '0.0"%"'
            else:
                vc = col_idx_to_var_cat.get(j)
                is_low = (unweighted_n_lookup and vc in unweighted_n_lookup
                          and unweighted_n_lookup[vc] < LOW_N_THRESHOLD)
                if is_low:
                    cell.fill = _pf("FFF8F0"); cell.font = Font(color="AA6020",size=9,name="Calibri")
                else:
                    bg_h,fg_h = _heat_fill(val)
                    cell.fill = _pf(bg_h); cell.font = Font(color=fg_h,size=9,name="Calibri")
                cell.alignment = _AC
                if isinstance(val,(int,float,np.floating)) and val == val:
                    cell.number_format = '0.0"%"'
        r += 1

    # ── NET rows ───────────────────────────────────────────────────────────────
    if approve_keys and disapprove_keys:
        r = _write_net_rows(ws, r, block_df, col_groups,
                            approve_keys, disapprove_keys,
                            col_idx_to_var_cat, group_starts)
    elif rep_keys and dem_keys:
        r = _write_horserace_net_rows(ws, r, block_df, group_starts, rep_keys, dem_keys)
    elif more_keys and less_keys:
        r = _write_morelikely_net_rows(ws, r, block_df, group_starts, more_keys, less_keys)

    return r + 3

def build_q7_ranking_block(df_in, w_col, group_order, q7_rank_cols):
    """
    Returns (mean_df, rank1_df, col_groups).
    mean_df  — weighted mean rank per subgroup (lower = more important), sorted asc.
    rank1_df — % of respondents who ranked this issue #1, sorted desc by total %.
    Both share the same col_groups structure.
    """
    flat = [(gl,v,c) for gl,v,cats in group_order for c in cats]
    w = pd.to_numeric(df_in[w_col], errors="coerce").fillna(0.0).to_numpy()
    var_cat_masks = {}
    for _,var,cat in flat:
        if (var,cat) not in var_cat_masks and var in df_in.columns:
            var_cat_masks[(var,cat)] = (df_in[var].astype(str).str.strip() == str(cat)).to_numpy()

    # Composite column keys to avoid category-name collisions
    seen_keys: dict = {}
    unique_col_keys = []
    for _,var,cat in flat:
        k = f"{var}|||{cat}"
        if k in seen_keys:
            seen_keys[k] += 1
            unique_col_keys.append(f"{k}||#{seen_keys[k]}")
        else:
            seen_keys[k] = 0
            unique_col_keys.append(k)

    mean_rows  = []
    rank1_rows = []

    for issue, col in q7_rank_cols.items():
        mean_rv  = [issue]
        rank1_rv = [issue]
        if col in df_in.columns:
            rank_arr = pd.to_numeric(df_in[col], errors="coerce").to_numpy()
            valid    = ~np.isnan(rank_arr)
            is_rank1 = (rank_arr == 1).astype(float)

            denom_total = float(w[valid].sum())
            mean_total  = (float((rank_arr[valid]*w[valid]).sum())/denom_total
                           if denom_total > 0 else np.nan)
            denom_r1    = float(w.sum())
            rank1_total = (float((is_rank1*w).sum())/denom_r1*100
                           if denom_r1 > 0 else np.nan)

            mean_rv.append(mean_total)
            rank1_rv.append(rank1_total)

            for _,var,cat in flat:
                vc = (var,cat)
                if vc in var_cat_masks:
                    m  = var_cat_masks[vc]
                    mv = m & valid
                    dv = float(w[mv].sum())
                    mean_rv.append(float((rank_arr[mv]*w[mv]).sum())/dv if dv>0 else np.nan)
                    dr1 = float(w[m].sum())
                    rank1_rv.append(float((is_rank1[m]*w[m]).sum())/dr1*100 if dr1>0 else np.nan)
                else:
                    mean_rv.append(np.nan)
                    rank1_rv.append(np.nan)
        else:
            mean_rv.extend([np.nan]*(1+len(flat)))
            rank1_rv.extend([np.nan]*(1+len(flat)))

        mean_rows.append(mean_rv)
        rank1_rows.append(rank1_rv)

    mean_df = (pd.DataFrame(mean_rows,  columns=["Issue","Total (Avg Rank)"] + unique_col_keys)
               .sort_values("Total (Avg Rank)", ascending=True).reset_index(drop=True))

    rank1_df_raw = pd.DataFrame(rank1_rows, columns=["Issue","Total (% Ranked #1)"] + unique_col_keys)
    # Reorder rank1 to match mean issue order, then sort by total % desc for display
    issue_order = mean_df["Issue"].tolist()
    rank1_df = (rank1_df_raw.set_index("Issue").reindex(issue_order)
                .reset_index().rename(columns={"index":"Issue"}))

    col_groups = []
    cur = 3
    for gl,var,cats in group_order:
        col_groups.append((gl, cur, cur+len(cats)-1))
        cur += len(cats)
    return mean_df, rank1_df, col_groups

def write_q7_block(ws, start_row, mean_df_or_tuple, col_groups_or_none=None):
    """
    Accepts either:
      write_q7_block(ws, row, mean_df, col_groups)   ← old 4-arg form (mean only)
      write_q7_block(ws, row, (mean_df, rank1_df, col_groups))  ← 3-arg tuple form
    Always writes BOTH the mean-rank table AND the % Ranked #1 sub-section.
    """
    # Handle both call signatures
    if isinstance(mean_df_or_tuple, tuple):
        mean_df, rank1_df, col_groups = mean_df_or_tuple
    else:
        mean_df = mean_df_or_tuple
        col_groups = col_groups_or_none
        rank1_df = None  # No rank1 data in legacy call

    _RH = [(8.5,"F7F9FC","1B2A3B"),(7.0,"D6E8F5","1B2A3B"),(5.5,"9DC3E6","1B2A3B"),
           (4.0,"5B9BD5","FFFFFF"),(2.5,"2E6EB0","FFFFFF"),(1.0,"1A4F8A","FFFFFF")]
    def _rh(v):
        try: v=float(v)
        except: return (_C_ROW_ODD,_C_DATA_FG)
        for thr,bg,fg in _RH:
            if v>=thr: return (bg,fg)
        return ("1A4F8A","FFFFFF")

    ncols = mean_df.shape[1]
    gs    = {sc for _,sc,_ in col_groups}
    r     = start_row

    # ── SECTION 1: Mean weighted rank ────────────────────────────────────────
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)
    c = ws.cell(r,1,value="Q7 Rank these 10 issues from most important to least important "
                           "(Mean weighted rank; 1 = most important — sorted best to worst)")
    c.fill=_pf(_C_TITLE_BG); c.font=Font(color=_C_TITLE_FG,bold=True,size=11,name="Calibri")
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=False,indent=1)
    ws.row_dimensions[r].height=_H_TITLE; r+=1

    # Group header
    ws.row_dimensions[r].height=_H_GROUP
    ws.cell(r,1).fill=_pf(_C_GRP_BG); ws.cell(r,1).border=_bd()
    tc=ws.cell(r,2,value="Total (Avg)")
    tc.fill=_pf(_C_TOTAL_BG); tc.font=Font(color=_C_TOTAL_FG,bold=True,size=9,name="Calibri")
    tc.alignment=_AC; tc.border=_bd()
    for gi,(gl,sc,ec) in enumerate(col_groups):
        bg=_GROUP_ACCENTS[gi%len(_GROUP_ACCENTS)]
        if sc<ec: ws.merge_cells(start_row=r,start_column=sc,end_row=r,end_column=ec)
        cell=ws.cell(r,sc,value=str(gl))
        cell.fill=_pf(bg); cell.font=Font(color="FFFFFF",bold=True,size=9,name="Calibri")
        cell.alignment=_AC
        for cc in range(sc,ec+1):
            ws.cell(r,cc).fill=_pf(bg)
            ws.cell(r,cc).border=_left_bd() if cc==sc else _bd()
    r+=1

    # Sub-header
    ws.row_dimensions[r].height=_H_SUBHDR
    for j,cn in enumerate(mean_df.columns,start=1):
        display_cn = str(cn).split("|||")[-1].split("||#")[0] if "|||" in str(cn) else str(cn)
        cell=ws.cell(r,j,value=display_cn)
        cell.fill=_pf(_C_SUB_BG); cell.font=Font(color=_C_SUB_FG,bold=False,size=8,name="Calibri")
        cell.alignment=_AC if j>1 else Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1)
        cell.border=_left_bd() if j in gs else _bd()
    r+=1

    # Data rows
    for i in range(mean_df.shape[0]):
        ws.row_dimensions[r].height=_H_DATA
        row_bg=_C_ROW_ODD if i%2==0 else _C_ROW_EVEN
        for j in range(1,ncols+1):
            val=mean_df.iloc[i,j-1]; cell=ws.cell(r,j,value=val)
            cell.border=_left_bd() if j in gs else _bd()
            if j==1:
                cell.fill=_pf(row_bg); cell.font=Font(color=_C_LABEL_FG,size=9,name="Calibri")
                cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=False,indent=1)
            elif j==2:
                cell.fill=_pf(_C_TOTAL_COL); cell.font=Font(color=_C_DATA_FG,bold=True,size=9,name="Calibri")
                cell.alignment=_AC
                if isinstance(val,(int,float,np.floating)) and val==val: cell.number_format='0.00'
            else:
                bg_h,fg_h=_rh(val); cell.fill=_pf(bg_h); cell.font=Font(color=fg_h,size=9,name="Calibri")
                cell.alignment=_AC
                if isinstance(val,(int,float,np.floating)) and val==val: cell.number_format='0.00'
        r+=1

    if rank1_df is None:
        return r+3

    # ── SECTION 2: % Ranked #1 ───────────────────────────────────────────────
    r += 2  # spacer
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)
    c2=ws.cell(r,1,value="Q7 Issues ranked #1 by respondents (% who placed each issue first — "
                          "shows each voter's single top priority)")
    c2.fill=_pf("163352"); c2.font=Font(color=_C_TITLE_FG,bold=True,size=11,name="Calibri")
    c2.alignment=Alignment(horizontal="left",vertical="center",wrap_text=False,indent=1)
    ws.row_dimensions[r].height=_H_TITLE; r+=1

    # Group header
    ws.row_dimensions[r].height=_H_GROUP
    ws.cell(r,1).fill=_pf(_C_GRP_BG); ws.cell(r,1).border=_bd()
    tc2=ws.cell(r,2,value="Total (%)")
    tc2.fill=_pf(_C_TOTAL_BG); tc2.font=Font(color=_C_TOTAL_FG,bold=True,size=9,name="Calibri")
    tc2.alignment=_AC; tc2.border=_bd()
    for gi,(gl,sc,ec) in enumerate(col_groups):
        bg=_GROUP_ACCENTS[gi%len(_GROUP_ACCENTS)]
        if sc<ec: ws.merge_cells(start_row=r,start_column=sc,end_row=r,end_column=ec)
        cell=ws.cell(r,sc,value=str(gl))
        cell.fill=_pf(bg); cell.font=Font(color="FFFFFF",bold=True,size=9,name="Calibri")
        cell.alignment=_AC
        for cc in range(sc,ec+1):
            ws.cell(r,cc).fill=_pf(bg)
            ws.cell(r,cc).border=_left_bd() if cc==sc else _bd()
    r+=1

    # Sub-header
    ws.row_dimensions[r].height=_H_SUBHDR
    for j,cn in enumerate(rank1_df.columns,start=1):
        display_cn = str(cn).split("|||")[-1].split("||#")[0] if "|||" in str(cn) else str(cn)
        cell=ws.cell(r,j,value=display_cn)
        cell.fill=_pf(_C_SUB_BG); cell.font=Font(color=_C_SUB_FG,bold=False,size=8,name="Calibri")
        cell.alignment=_AC if j>1 else Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1)
        cell.border=_left_bd() if j in gs else _bd()
    r+=1

    # Sort rank1_df descending by total % for display
    rank1_sorted = rank1_df.sort_values("Total (% Ranked #1)", ascending=False).reset_index(drop=True)
    for i in range(rank1_sorted.shape[0]):
        ws.row_dimensions[r].height=_H_DATA
        row_bg=_C_ROW_ODD if i%2==0 else _C_ROW_EVEN
        for j in range(1,ncols+1):
            val=rank1_sorted.iloc[i,j-1]; cell=ws.cell(r,j,value=val)
            cell.border=_left_bd() if j in gs else _bd()
            if j==1:
                cell.fill=_pf(row_bg); cell.font=Font(color=_C_LABEL_FG,size=9,name="Calibri")
                cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=False,indent=1)
            elif j==2:
                cell.fill=_pf(_C_TOTAL_COL); cell.font=Font(color=_C_DATA_FG,bold=True,size=9,name="Calibri")
                cell.alignment=_AC
                if isinstance(val,(int,float,np.floating)) and val==val: cell.number_format='0.0"%"'
            else:
                bg_h,fg_h=_heat_fill(val); cell.fill=_pf(bg_h); cell.font=Font(color=fg_h,size=9,name="Calibri")
                cell.alignment=_AC
                if isinstance(val,(int,float,np.floating)) and val==val: cell.number_format='0.0"%"'
        r+=1
    return r+3

def build_q2_multiselect_block(df_in, w_col, group_order):
    flat = [(gl,v,c) for gl,v,cats in group_order for c in cats]
    w    = pd.to_numeric(df_in[w_col],errors="coerce").fillna(0.0).astype(float)
    wt   = float(w.sum())
    varcw = {}
    for _,var,_ in flat:
        if var not in varcw and var in df_in.columns:
            varcw[var] = {cat: float(w[df_in[var]==cat].sum())
                          for cat in df_in[var].astype(str).str.strip().unique()}
    # Composite column keys to avoid category-name collisions
    seen_keys: dict = {}
    unique_col_keys = []
    for _,var,cat in flat:
        k = f"{var}|||{cat}"
        if k in seen_keys:
            seen_keys[k] += 1
            unique_col_keys.append(f"{k}||#{seen_keys[k]}")
        else:
            seen_keys[k] = 0
            unique_col_keys.append(k)
    data = {"Response":[], "Total":[]}
    for ck in unique_col_keys: data[ck] = []
    for col_key,label in Q2_ITEMS_DISPLAY:
        data["Response"].append(label)
        if col_key in df_in.columns:
            x = pd.to_numeric(df_in[col_key],errors="coerce").fillna(0.0).astype(float)
            data["Total"].append(float((x*w).sum())/wt*100.0 if wt>0 else 0.0)
            for ck,(_,var,cat) in zip(unique_col_keys,flat):
                if var in df_in.columns:
                    mask = df_in[var]==cat; d = varcw[var].get(cat,0.0)
                    data[ck].append(float((x[mask]*w[mask]).sum())/d*100.0 if d>0 else 0.0)
                else: data[ck].append(0.0)
        else:
            data["Total"].append(0.0)
            for ck in unique_col_keys: data[ck].append(0.0)
    block_df = pd.DataFrame(data)
    col_groups=[]; cur=3
    for gl,_,cats in group_order:
        col_groups.append((gl,cur,cur+len(cats)-1)); cur+=len(cats)
    return block_df, col_groups

def find_matrix_columns(df_in, prefix):
    return [c for c in df_in.columns if isinstance(c,str) and prefix.lower() in c.lower()]

def _detect_response_order(df_in, cols, fallback):
    all_vals = set()
    for col in cols:
        if col in df_in.columns:
            vals = df_in[col].dropna().astype(str).str.strip().unique()
            all_vals.update(v for v in vals if v not in ("nan","","None"))
    ordered = [v for v in fallback if v in all_vals]
    extras  = sorted(v for v in all_vals if v not in fallback)
    return (ordered+extras) if (ordered+extras) else fallback

def write_matrix_into_tabbook(ws, start_row, df_in, weight_col, prefix, section_title,
                               unweighted_n_lookup=None):
    cols = find_matrix_columns(df_in, prefix)
    row  = start_row
    if not cols:
        ws.cell(row,1,value=f"(No columns found for {section_title} — prefix: {prefix!r})")
        return row+2
    low = section_title.lower()
    expected = APPROVE_RESPONSE_ORDER
    resp_order = _detect_response_order(df_in, cols, expected)
    context = section_title.split("(Matrix)")[0].strip()
    for col in cols:
        item = str(col)
        if prefix.lower() in item.lower():
            item = item[item.lower().index(prefix.lower())+len(prefix):].strip()
        item = item.lstrip("—-: ").strip()
        item_title = f"{context} — {item}"
        bd,cg = build_block_table(df_in, col, weight_col, GROUP_ORDER, resp_order)
        row = write_block_to_sheet(ws, row, item_title, bd, cg, unweighted_n_lookup)
    return row

# ==============================================================================
# 13) TABBOOK SEQUENCE — all questions this poll contains
# ==============================================================================

TABBOOK_SEQUENCE = [
    ("single",    "Q3_VoteIntent"),
    ("single",    "Q4_BallotMethod"),
    ("single",    "Q5_SocialVote"),
    ("single",    "Q6_2024Vote"),
    ("q2",        None),
    ("q7_ranking",None),
    ("single",    "Q8_PoliticalOutlook"),
    ("single",    "Q9_PartyID"),
    ("single",    "Q10_Groyper"),
    ("single",    "Q11_RightTrack"),
    ("single",    "Q13_GenericBallot"),
    ("single",    "Q14_TrumpApprove"),
    ("matrix",    ("Q12_2028_Matchups",  "Q12 2028 Presidential Matchups (Matrix)")),
    ("matrix",    ("Q16_IssueHandling",  "Q16 Trump Issue Handling (Matrix)")),
    ("matrix",    ("Q17_IndivApproval",  "Q17 Individual Approval (Matrix)")),
    ("single",    "Q18_TrumpConservatism"),
    ("single",    "Q19_ForeignPolicy"),
    ("single",    "Q20_HHExpenses"),
    ("single",    "Q21_MassDeportation"),
    ("single",    "Q22_GenPref"),
    ("single",    "Q23_PartnerQuality"),
    ("single",    "Q24_IsraelPAC"),
    ("single",    "Q25_CharlieKirk"),
]

def write_tabbook_sheet(writer, sheet_name, df_in, weight_col, q7_rank_cols,
                         resolved_matrix):
    global _CURRENT_GROUP_ORDER_REF
    _CURRENT_GROUP_ORDER_REF = GROUP_ORDER
    ws  = writer.book.create_sheet(sheet_name)
    row = 1
    ws.cell(row,  1, value="Survey name:")
    ws.cell(row+1,1, value=SURVEY_NAME)
    row += 3

    _unwt = build_unweighted_n_lookup(df_in, GROUP_ORDER)

    for col_key,_ in Q2_ITEMS_DISPLAY:
        if col_key in df_in.columns:
            df_in[col_key] = pd.to_numeric(df_in[col_key],errors="coerce").fillna(0).astype(float)

    # Net-row config for approve/disapprove questions
    APPROVE_QUESTION_COLS = {
        "Q14_TrumpApprove": {
            "approve_keys":    ["Strongly approve", "Somewhat approve"],
            "disapprove_keys": ["Strongly disapprove", "Somewhat disapprove"],
        },
    }
    # More-likely / less-likely
    MORELIKELY_QUESTION_COLS = {
        "Q24_IsraelPAC": {
            "more_keys": ["Much more likely", "Somewhat more likely"],
            "less_keys":  ["Much less likely", "Somewhat less likely"],
        },
    }
    # Horse-race questions
    HORSERACE_QUESTION_COLS = {"Q13_GenericBallot"}

    total_items = len(TABBOOK_SEQUENCE)
    for seq_idx,(kind,payload) in enumerate(TABBOOK_SEQUENCE,1):
        lbl = payload if isinstance(payload,str) else (payload[1] if isinstance(payload,tuple) else kind)
        print(f"  [{seq_idx:>2}/{total_items}] {sheet_name} — {lbl}")

        if kind == "single":
            q = payload
            if q not in df_in.columns:
                print(f"    ⚠️  Column not found: {q}"); continue
            meta   = QUESTION_ORDER_SINGLE.get(q,{})
            title  = meta.get("title",q)
            opts   = meta.get("options",
                     sorted(df_in[q].dropna().astype(str).unique().tolist()))
            actual_vals = set(df_in[q].dropna().astype(str).str.strip().unique()) - {"nan",""}
            ordered = [o for o in opts if o in actual_vals]
            ordered += sorted(v for v in actual_vals if v not in opts)
            if not ordered: ordered = opts
            bd,cg  = build_block_table(df_in,q,weight_col,GROUP_ORDER,ordered)

            a_keys = d_keys = r_keys = dm_keys = mk_keys = lk_keys = None
            net_cfg = APPROVE_QUESTION_COLS.get(q)
            if net_cfg:
                a_keys = net_cfg["approve_keys"]; d_keys = net_cfg["disapprove_keys"]
            elif q in HORSERACE_QUESTION_COLS:
                r_keys, dm_keys = detect_horserace_keys(opts)
            elif q in MORELIKELY_QUESTION_COLS:
                ml_cfg = MORELIKELY_QUESTION_COLS[q]
                mk_keys = ml_cfg["more_keys"]; lk_keys = ml_cfg["less_keys"]

            row = write_block_to_sheet(ws,row,title,bd,cg,_unwt,
                                       approve_keys=a_keys, disapprove_keys=d_keys,
                                       rep_keys=r_keys, dem_keys=dm_keys,
                                       more_keys=mk_keys, less_keys=lk_keys); continue

        if kind == "q2":
            bd,cg = build_q2_multiselect_block(df_in,weight_col,GROUP_ORDER)
            row   = write_block_to_sheet(ws,row,
                "Q2 Which of the following election years did you vote in at least once? (% selecting)",
                bd,cg,_unwt); continue

        if kind == "q7_ranking":
            mean_df, rank1_df, cg = build_q7_ranking_block(df_in,weight_col,GROUP_ORDER,q7_rank_cols)
            row = write_q7_block(ws, row, (mean_df, rank1_df, cg)); continue

        if kind == "matrix":
            prefix_key, section_title = payload
            actual_prefix = resolved_matrix.get(prefix_key)
            if not actual_prefix:
                print(f"    ⚠️  Matrix prefix not resolved for {prefix_key}"); continue
            row = write_matrix_into_tabbook(ws,row,df_in,weight_col,
                                             actual_prefix,section_title,_unwt); continue

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 11
    for ci in range(3,220):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "C4"
    ws.sheet_properties.tabColor = "0D1B2A"

# ==============================================================================
# 14) ELECTORATE COMPOSITION SHEET
# ==============================================================================

def composition_table(df_in, wcol, var, order=None):
    if len(df_in)==0 or var not in df_in.columns: return pd.DataFrame()
    s  = df_in[var].astype(str).str.strip().replace({"nan":np.nan,"":np.nan,"None":np.nan})
    w  = pd.to_numeric(df_in[wcol],errors="coerce").fillna(0.0)
    t  = pd.DataFrame({"cat":s,"w":w}).dropna(subset=["cat"])
    wn = t.groupby("cat")["w"].sum(); base = float(wn.sum())
    pct= (wn/base*100.0) if base>0 else wn*0.0
    out= pd.DataFrame({"Variable":var,"Category":wn.index,"Weighted_N":wn.values,"Weighted_%":pct.values})
    if order:
        out["__o__"] = out["Category"].map({k:i for i,k in enumerate(order)}).fillna(9999)
        out = out.sort_values(["__o__","Category"]).drop(columns="__o__")
    else:
        out = out.sort_values("Weighted_N",ascending=False)
    return out.reset_index(drop=True)

def write_composition_block(ws, start_row, title, comp_df):
    cols=["Variable","Category","Weighted_N","Weighted_%"]; r=start_row
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=len(cols))
    c = ws.cell(r,1,value=title)
    c.fill=_pf(_C_TITLE_BG); c.font=Font(color=_C_TITLE_FG,bold=True,size=11,name="Calibri")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[r].height=_H_TITLE; r+=1
    ws.row_dimensions[r].height=_H_SUBHDR
    for j,col in enumerate(cols,start=1):
        cell=ws.cell(r,j,value=col)
        cell.fill=_pf(_C_SUB_BG); cell.font=Font(color=_C_SUB_FG,bold=True,size=9,name="Calibri")
        cell.alignment=_AC; cell.border=_bd()
    prev_var=None
    for idx,(_, row_data) in enumerate(comp_df.iterrows()):
        r+=1; ws.row_dimensions[r].height=_H_DATA
        row_bg=_C_ROW_ODD if idx%2==0 else _C_ROW_EVEN
        cur_var=row_data["Variable"]
        c1=ws.cell(r,1,value=str(cur_var) if cur_var!=prev_var else "")
        c1.fill=_pf("E8EFF6"); c1.font=Font(color="163352",bold=True,size=9,name="Calibri")
        c1.alignment=Alignment(horizontal="left",vertical="center",indent=1); c1.border=_bd()
        prev_var=cur_var
        c2=ws.cell(r,2,value=str(row_data["Category"]))
        c2.fill=_pf(row_bg); c2.font=Font(color=_C_LABEL_FG,size=9,name="Calibri")
        c2.alignment=Alignment(horizontal="left",vertical="center",indent=1); c2.border=_bd()
        c3=ws.cell(r,3,value=float(row_data["Weighted_N"]))
        c3.fill=_pf(_C_TOTAL_COL); c3.font=Font(color=_C_DATA_FG,size=9,name="Calibri")
        c3.alignment=_AC; c3.number_format="0.0"; c3.border=_bd()
        pv=float(row_data["Weighted_%"]); bg_h,fg_h=_heat_fill(pv)
        c4=ws.cell(r,4,value=pv)
        c4.fill=_pf(bg_h); c4.font=Font(color=fg_h,size=9,name="Calibri")
        c4.alignment=_AC; c4.number_format='0.0"%"'; c4.border=_bd()
    return r+3

def build_electorate_sheet(writer, sheet_name, rv_df, lv_df):
    ws = writer.book.create_sheet(sheet_name)
    vars_to_show = [
        ("Age",             ["18-29","30-44","45-64","65+"]),
        ("Gender",          ["Male","Female"]),
        ("AgeGender",       list(AGE_GENDER_TARGETS.keys())),
        ("RaceEdu",         ["White No College","White College","Hispanic","Black","Asian / Other"]),
        ("Education4",      ["High school or less","Some college/assoc. degree",
                             "College graduate","Postgraduate study"]),
        ("Income",          ["$0–$25k","$25–$50k","$50–$75k","$75–$100k","$100–$150k","$150–$200k","$200k+"]),
        ("Region",          NATIONAL_REGIONS),
        ("Party",           ["Republican","Democrat","Independent"]),
        ("Party_Detailed",  PARTY_DETAILED_ORDER_DISPLAY),
        ("Vote2024_Bucket", ["Donald Trump","Kamala Harris","Third party","Did not vote"]),
        ("VoteHistory",     ["Consistent voter","Occasional voter","New / non-voter"]),
        ("Groyper_Bucket",  ["Yes", "No"]),
    ]
    row = 1
    for block_title,dfi,wcol in [
        ("Registered Voters Electorate (weighted)",      rv_df, "weight_rv"),
        ("Likely Voters Electorate (derived; weighted)", lv_df, "weight_lv"),
    ]:
        all_comp = [composition_table(dfi,wcol,var,order=order) for var,order in vars_to_show]
        comp_df  = pd.concat([x for x in all_comp if len(x)>0], ignore_index=True)
        row = write_composition_block(ws,row,block_title,comp_df)
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=60
    ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=14
    ws.sheet_properties.tabColor="0D1B2A"

# ==============================================================================
# 15) TERMINAL DIAGNOSTICS (toplines to stdout + file)
# ==============================================================================

def _print_weighted_dist(label, df_in, col, wcol, ordered_responses=None):
    if col not in df_in.columns: print(f"  ⚠️  Column not found: {col}"); return
    s = df_in[col].astype(str).str.strip().replace({"nan":None,"":None,"None":None})
    w = pd.to_numeric(df_in[wcol],errors="coerce").fillna(0.0)
    wn = {}
    for resp,wv in zip(s[s.notna()],w[s.notna()]): wn[resp]=wn.get(resp,0.0)+float(wv)
    tot=sum(wn.values())
    if tot==0: print(f"  ⚠️  No weighted data for {col}"); return
    keys = list(ordered_responses or []) if ordered_responses else []
    keys = [r for r in keys if r in wn] + [r for r in sorted(wn) if r not in keys]
    print(f"\n  {'Response':<55}  {'Wtd %':>7}  {'Wtd N':>8}")
    print(f"  {'-'*55}  {'-'*7}  {'-'*8}")
    for resp in keys:
        pct = wn.get(resp,0.0)/tot*100.0
        for il,line in enumerate(textwrap.wrap(str(resp),width=53),0):
            if il==0: print(f"  {line:<55}  {pct:>6.1f}%  {wn.get(resp,0.0):>8.1f}")
            else:     print(f"    {line}")
    print(f"\n  Base (unweighted n): {int(s.notna().sum()):,}  |  Total weighted N: {tot:,.1f}")

def print_all_questions(rv_df, lv_df, q7_rank_cols, weight_rv="weight_rv", weight_lv="weight_lv"):
    def _divider(c="=",w=80): print(c*w)
    def _both(label,col,opts=None):
        _divider("-")
        print(f"\n  {label}")
        print(f"\n  ── Registered Voters (weighted) ──")
        _print_weighted_dist(label,rv_df,col,weight_rv,opts)
        print(f"\n  ── Likely Voters (weighted) ──")
        _print_weighted_dist(label,lv_df,col,weight_lv,opts)

    _divider("=")
    print(f"  WEIGHTED TOPLINES — {SURVEY_NAME}")
    print(f"  RV n={len(rv_df):,}  |  LV n={len(lv_df):,}")
    _divider("=")

    for qcol,meta in QUESTION_ORDER_SINGLE.items():
        _both(meta["title"], qcol, meta.get("options"))

    _divider("-")
    print("\n  Q7 Rank these 10 issues (Mean weighted rank; 1 = most important)")
    for df_lbl,dfi,wcol in [("Registered Voters",rv_df,weight_rv),("Likely Voters",lv_df,weight_lv)]:
        print(f"\n  ── {df_lbl} ──")
        w = pd.to_numeric(dfi[wcol],errors="coerce").fillna(0.0).to_numpy()
        results=[]
        for issue,col in q7_rank_cols.items():
            if col in dfi.columns:
                ra=pd.to_numeric(dfi[col],errors="coerce").to_numpy(); ok=~np.isnan(ra)
                d=float(w[ok].sum())
                results.append((float((ra[ok]*w[ok]).sum())/d if d>0 else np.nan, issue))
            else: results.append((np.nan,issue))
        results.sort(key=lambda x: x[0] if not np.isnan(x[0]) else 999)
        print(f"\n  {'#':>4}  {'Issue':<52}  {'Avg Rank':>9}")
        print(f"  {'-'*4}  {'-'*52}  {'-'*9}")
        for rp,(avg,issue) in enumerate(results,1):
            print(f"  {'#'+str(rp):>4}  {issue:<52}  {avg:>9.2f}")
    _divider("=")

# ==============================================================================
# 16) MAIN
# ==============================================================================

def main(csv_path=None):
    t0 = _time.time()
    print(f"\n{'='*60}")
    print(f"  {SURVEY_NAME}")
    print(f"  Weighting Script — Wisconsin/US Weights Methodology")
    print(f"{'='*60}\n")

    # ── Load ──────────────────────────────────────────────────────────────────
    if csv_path is None:
        csv_path = INPUT_FILE
    if not os.path.exists(csv_path):
        # Also try same directory as script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        csv_path = os.path.join(script_dir, os.path.basename(csv_path))
    print(f"Loading CSV: {csv_path}")
    df_raw = pd.read_csv(csv_path, low_memory=False)
    print(f"  Loaded {len(df_raw):,} rows × {len(df_raw.columns)} columns")

    # ── Auto-detect columns ──────────────────────────────────────────────────
    df, resolved_matrix, q7_rank_cols, q2_cols = auto_rename_columns(df_raw)

    # ── Demographics ──────────────────────────────────────────────────────────
    print("\nBuilding demographic variables...")

    # Age — from Pollfish TPSI Age column (single-file: actual numeric age)
    age_raw_col = next((c for c in ["TPSI Age","Age","age"] if c in df.columns), None)
    if age_raw_col:
        def _age_bucket(v):
            try: a = int(float(str(v).strip()))
            except: return "45-64"
            if a <= 29: return "18-29"
            if a <= 44: return "30-44"
            if a <= 64: return "45-64"
            return "65+"
        df["Age"] = df[age_raw_col].apply(_age_bucket)
    elif "_cohort" in df.columns:
        df["Age"] = df["_cohort"].astype(str).str.strip()
    else:
        print("  ⚠️  Age column not found — defaulting to '45-64'")
        df["Age"] = "45-64"

    # Gender
    gcol = next((c for c in ["DEMO_Gender","gender","Gender","TPSI Gender"]
                 if c in df.columns), None)
    if gcol:
        df["Gender"] = df[gcol].apply(normalize_gender)
    else:
        print("  ⚠️  Gender column not found — defaulting to 'Unknown'")
        df["Gender"] = "Unknown"

    # Age × Gender joint cell
    df["AgeGender"] = df["Age"] + "_" + df["Gender"]

    # Race / Education
    rcol = next((c for c in ["DEMO_Race","race","Race","TPSI Race"] if c in df.columns), None)
    ecol = next((c for c in ["DEMO_Education","education","Education","TPSI Education"] if c in df.columns), None)
    excol= next((c for c in ["DEMO_Ethnicity","ethnicity","Ethnicity","TPSI Ethnicity"] if c in df.columns), None)

    if rcol and ecol:
        df["RaceEdu"] = df.apply(
            lambda r: normalize_race_edu(
                str(r.get(rcol,"")),
                str(r.get(ecol,"")),
                str(r.get(excol,"")) if excol else ""
            ), axis=1
        )
    else:
        print("  ⚠️  Race/Education columns not found — RaceEdu dimension will have limited coverage")
        df["RaceEdu"] = "White No College"

    # Education 4-way
    if ecol:
        df["Education4"] = df[ecol].apply(normalize_education_4way)
    else:
        df["Education4"] = "High school or less"

    # Education binary + Gender × Education joint cell (for raking)
    if ecol:
        df["Education"] = df[ecol].apply(normalize_education_binary)
    else:
        df["Education"] = "No College"

    gcol_raw = next((c for c in ["DEMO_Gender","gender","Gender","TPSI Gender"]
                     if c in df.columns), None)
    if gcol_raw and ecol:
        df["GenderEdu"] = df.apply(
            lambda r: normalize_gender_edu(str(r.get(gcol_raw,"")), str(r.get(ecol,""))), axis=1
        )
    else:
        df["GenderEdu"] = "Female_No College"

    # Region (from state column)
    scol = next((c for c in ["DEMO_State","TPSI State","TPSI Region","State"] if c in df.columns), None)
    if scol:
        df["Region"] = df[scol].apply(normalize_state_to_region)
    else:
        print("  ⚠️  State column not found — Region dimension will be skipped")
        df["Region"] = "Southeast Atlantic"

    # Party (from Q8 political outlook)
    if "Q8_PoliticalOutlook" in df.columns:
        df["Party"] = df["Q8_PoliticalOutlook"].apply(party_from_q8)
        df["Party_Detailed"] = df["Q8_PoliticalOutlook"].apply(
            lambda x: _party_display_map.get(str(x).strip(), str(x)[:40])
        )
    elif "Q9_PartyID" in df.columns:
        def _pid(v):
            v=str(v).strip().lower()
            if "democrat" in v: return "Democrat"
            if "republican" in v: return "Republican"
            return "Independent"
        df["Party"] = df["Q9_PartyID"].apply(_pid)
        df["Party_Detailed"] = df["Party"]
    else:
        df["Party"] = "Independent"; df["Party_Detailed"] = "Independent"

    # 2024 Vote bucket
    if "Q6_2024Vote" in df.columns:
        df["Vote2024_Bucket"] = df["Q6_2024Vote"].apply(vote2024_bucket)
    else:
        df["Vote2024_Bucket"] = np.nan

    # Vote history bucket — pass q2_cols dict which maps year → column name
    df["VoteHistory"] = df.apply(lambda r: make_vote_history_bucket(r, q2_cols), axis=1)

    # Groyper awareness bucket (Q10 → short-label alias for crosstab columns)
    def _remap_col(df_in, src_col, label_map, new_col):
        """Map full response labels to short display labels for a new column."""
        if src_col not in df_in.columns:
            df_in[new_col] = np.nan
            return df_in
        df_in[new_col] = df_in[src_col].astype(str).str.strip().map(label_map)
        return df_in

    # Groyper identity (Q10) — Yes/No binary crosstab column
    # The raw column already contains "Yes" / "No"; just clean and alias it.
    if "Q10_Groyper" in df.columns:
        df["Groyper_Bucket"] = (df["Q10_Groyper"].astype(str).str.strip()
                                .map({"Yes": "Yes", "No": "No"}))
        n_yes = (df["Groyper_Bucket"] == "Yes").sum()
        n_no  = (df["Groyper_Bucket"] == "No").sum()
        print(f"  Groyper crosstab column derived: Yes={n_yes:,}  No={n_no:,}")
    else:
        df["Groyper_Bucket"] = np.nan
        print("  ⚠️  Q10_Groyper column not found — Groyper crosstab will be empty")

    # Income — derive from DEMO_Income column if present
    icol = next((c for c in ["DEMO_Income","TPSI Income","Household income US",
                              "household income","income"] if c in df.columns), None)
    if icol:
        df["Income"] = df[icol].apply(normalize_income)
        n_inc = df["Income"].notna().sum()
        print(f"  Income column derived from '{icol}': {n_inc:,} non-null values")
    else:
        df["Income"] = np.nan
        print("  ⚠️  Income column not found — Income crosstab will be empty")

    # Also ensure Q2_Never_Voted is present for vote history (check alternate names)
    if "Q2_Never_Voted" not in df.columns:
        for alt in [c for c in df.columns if "never" in c.lower() or
                    ("q2" in c.lower() and "elig" in c.lower())]:
            df["Q2_Never_Voted"] = df[alt]
            print(f"  Mapped '{alt}' → Q2_Never_Voted")
            break

    # Clean text columns
    for col in ["Q3_VoteIntent","Q4_BallotMethod","Q5_SocialVote","Q6_2024Vote",
                "Q8_PoliticalOutlook","Q9_PartyID","Q11_RightTrack","Q13_GenericBallot",
                "Q14_TrumpApprove"]:
        if col in df.columns:
            df[col] = clean_text_series(df[col])

    print(f"\n  Demographic summary (unweighted):")
    for var,cats in [("Age",["18-29","30-44","45-64","65+"]),
                     ("Gender",["Male","Female"]),
                     ("Party",["Republican","Democrat","Independent"]),
                     ("Vote2024_Bucket",["Donald Trump","Kamala Harris","Third party","Did not vote"])]:
        if var in df.columns:
            dist = df[var].value_counts(normalize=True)*100
            parts = [f"{c}: {dist.get(c,0):.1f}%" for c in cats]
            print(f"    {var:<18} {' | '.join(parts)}")

    # ── Build raking benchmark dictionary ─────────────────────────────────────
    benchmarks = {
        "AgeGender":      AGE_GENDER_TARGETS,
        "RaceEdu":        RACE_EDU_TARGETS,
        "Education4":     EDUCATION_TARGETS,
        "GenderEdu":      GENDER_EDU_TARGETS,
        "Region":         REGION_TARGETS,
        "Vote2024_Bucket":VOTE_2024_TARGETS,
    }
    # Drop dimensions with insufficient coverage (< 30%)
    active_benchmarks = {}
    for var, tgts in benchmarks.items():
        if var not in df.columns:
            print(f"  Skipping dimension '{var}' — column not found"); continue
        cov = df[var].isin(tgts.keys()).mean()
        if cov >= 0.30:
            active_benchmarks[var] = tgts
        else:
            print(f"  Skipping dimension '{var}' — only {cov:.1%} coverage")

    # ── Method 12: Cell collapse + Entropy Balancing + Raking ─────────────────
    RAKE_ROUNDS = 4; WEIGHT_CAP = 2.0
    rv_df = df.copy()

    print("  Checking for thin raking cells (n < 5)...")
    rv_collapsed, active_benchmarks_collapsed = collapse_thin_cells(rv_df, active_benchmarks, min_cell_n=5)

    # Method 5: Entropy balancing first, then iterative raking rounds with cap
    print("\n  Running Entropy Balancing (Method 5)...")
    rv_df = entropy_balance(rv_collapsed, active_benchmarks_collapsed, weight_col="weight_rv")
    if rv_df["weight_rv"].sum() > 0:
        rv_df["weight_rv"] *= len(rv_df) / rv_df["weight_rv"].sum()

    _nat_deff = compute_deff(rv_df["weight_rv"])
    print(f"  Natural DEFF (post-EB, no cap): {_nat_deff:.3f}")
    if _nat_deff < 1.3:   WEIGHT_CAP = 2.0
    elif _nat_deff < 1.8: WEIGHT_CAP = 1.75
    elif _nat_deff < 2.5: WEIGHT_CAP = 1.75
    else:                 WEIGHT_CAP = 1.75
    print(f"  → Setting weight cap to {WEIGHT_CAP}")

    for rnd in range(1, RAKE_ROUNDS+1):
        print(f"\n  ── Raking round {rnd}/{RAKE_ROUNDS} ──")
        rv_df = rake_weights(rv_df, active_benchmarks_collapsed, max_iter=80,
                             weight_col="weight_rv", init_weights=rv_df["weight_rv"], verbose=True)
        if rv_df["weight_rv"].sum() > 0:
            rv_df["weight_rv"] *= len(rv_df) / rv_df["weight_rv"].sum()

        n_capped = int((rv_df["weight_rv"] > WEIGHT_CAP).sum())
        rv_df["weight_rv"] = rv_df["weight_rv"].clip(upper=WEIGHT_CAP)
        if rv_df["weight_rv"].sum() > 0:
            rv_df["weight_rv"] *= len(rv_df) / rv_df["weight_rv"].sum()

        deff_cur = compute_deff(rv_df["weight_rv"])
        effn_cur = effective_n(rv_df["weight_rv"])
        wmin=rv_df["weight_rv"].min(); wmax=rv_df["weight_rv"].max()
        print(f"  Round {rnd}: min={wmin:.3f}  max={wmax:.3f}  "
              f"capped={n_capped}  DEFF={deff_cur:.3f}  Eff.N={effn_cur:.0f}")
        if n_capped == 0:
            print(f"  No weights capped — converged after round {rnd}"); break

    deff_final = compute_deff(rv_df["weight_rv"])
    effn_final = effective_n(rv_df["weight_rv"])
    print_deff_summary("RV weights post-raking", rv_df["weight_rv"])

    # ── Method 8: Response Propensity Scores ──────────────────────────────────
    print("\n  Computing response propensity scores (Method 8)...")
    _PROPENSITY_VARS = ["AgeGender", "Gender", "Education4", "RaceEdu"]
    ps = compute_propensity_scores(rv_df, active_benchmarks_collapsed, _PROPENSITY_VARS)
    rv_df["PropensityScore"] = ps
    n_low_ps = int((ps < 0.10).sum())
    if n_low_ps > 0:
        print(f"  ⚠️  {n_low_ps} respondents with propensity <0.10 (extreme non-responders)")

    # ── 2024 Recall Calibration ────────────────────────────────────────────────
    if rv_df["Vote2024_Bucket"].notna().sum() > 0:
        print("\n  Applying 2024 recall calibration (two-stage: FEC voters + CPS non-voter)...")
        rv_df = apply_recall2024_calibration(rv_df)
        n_recapped = int((rv_df["weight_rv"] > WEIGHT_CAP).sum())
        rv_df["weight_rv"] = rv_df["weight_rv"].clip(upper=WEIGHT_CAP)
        if rv_df["weight_rv"].sum() > 0:
            rv_df["weight_rv"] *= len(rv_df) / rv_df["weight_rv"].sum()
        print(f"  Post-calibration: re-capped={n_recapped}")
        print_deff_summary("RV weights post-recall-cal", rv_df["weight_rv"])

    # ── Likely Voter Model ─────────────────────────────────────────────────────
    rv_df["LV_Score"] = rv_df.apply(lambda r: compute_lv_score(r, q2_cols), axis=1)
    raw_lv = rv_df["weight_rv"] * rv_df["LV_Score"]
    lv_df  = rv_df.copy()
    if raw_lv.sum() > 0:
        lv_df["weight_lv"] = raw_lv * (len(lv_df) / raw_lv.sum())
    else:
        lv_df["weight_lv"] = lv_df["weight_rv"]

    deff_lv = compute_deff(lv_df["weight_lv"])
    print(f"\n  LV model: n={len(lv_df):,}  "
          f"LV score mean={rv_df['LV_Score'].mean():.4f}  "
          f"weight range: {lv_df['weight_lv'].min():.3f}–{lv_df['weight_lv'].max():.3f}")
    print_deff_summary("LV weights", lv_df["weight_lv"])

    # ── Method 10: ICC / Geographic Clustering ────────────────────────────────
    icc_results = {}
    if "Region" in rv_df.columns and "Vote2024_Bucket" in rv_df.columns:
        print("  Computing geographic clustering ICC (Method 10)...")
        icc_results = compute_icc(rv_df, "Region", "Vote2024_Bucket", "weight_rv")
        print(f"  ICC={icc_results.get('ICC','n/a')}  "
              f"DEFF_cluster={icc_results.get('DEFF_cluster','n/a')}")

    # ── Methods 9 + 11: Weight diagnostics & MoE ─────────────────────────────
    diag_rv    = weight_diagnostics_report(rv_df, "weight_rv", active_benchmarks_collapsed,
                                            len(rv_df), "PropensityScore")
    diag_lv    = weight_diagnostics_report(lv_df, "weight_lv", active_benchmarks_collapsed, len(lv_df))
    balance_rv = covariate_balance_table(rv_df, active_benchmarks_collapsed, "weight_rv")
    balance_lv = covariate_balance_table(lv_df, active_benchmarks_collapsed, "weight_lv")

    _BOOTSTRAP_QUESTIONS = ["Q6_2024Vote","Q3_VoteIntent","Q13_GenericBallot","Q14_TrumpApprove","Q11_RightTrack"]
    moe_rows = []
    for q in _BOOTSTRAP_QUESTIONS:
        if q not in rv_df.columns: continue
        s = rv_df[q].astype(str).str.strip()
        w = pd.to_numeric(rv_df["weight_rv"], errors="coerce").fillna(0)
        wn = {}
        for rsp, wv in zip(s, w):
            if rsp not in ("nan","","None"): wn[rsp] = wn.get(rsp, 0.0) + float(wv)
        tot = sum(wn.values())
        for resp, wv in sorted(wn.items(), key=lambda x: -x[1]):
            p   = wv / tot
            moe = compute_moe(p, len(rv_df), deff_final)
            moe_rows.append({"Question": q, "Response": resp,
                              "Weighted %": round(p*100, 1),
                              "MoE (±pp)": round(moe, 1) if not np.isnan(moe) else "n/a",
                              "DEFF used": round(deff_final, 3)})
    moe_df = pd.DataFrame(moe_rows)

    # ── Bootstrap SEs ──────────────────────────────────────────────────────────
    print(f"\n{'='*60}\n  BOOTSTRAP STANDARD ERRORS — KEY QUESTIONS (n_boot=500)\n{'='*60}")
    key_qs = [
        ("Q14_TrumpApprove",  "Strongly approve"),
        ("Q14_TrumpApprove",  "Strongly disapprove"),
        ("Q13_GenericBallot", "The Republican candidate"),
        ("Q13_GenericBallot", "The Democrat candidate"),
        ("Q11_RightTrack",    "Right track"),
        ("Q11_RightTrack",    "Wrong track"),
    ]
    print(f"  {'Question':<25}  {'Response':<30}  {'RV %':>6}  {'±SE':>5}  {'LV %':>6}  {'±SE':>5}")
    print(f"  {'-'*25}  {'-'*30}  {'-'*6}  {'-'*5}  {'-'*6}  {'-'*5}")
    for qcol,resp in key_qs:
        if qcol not in rv_df.columns: continue
        rw = pd.to_numeric(rv_df["weight_rv"],errors="coerce").fillna(0)
        rq = rv_df[qcol].astype(str).str.strip()
        rv_pct = rw[rq==resp].sum()/rw.sum()*100 if rw.sum()>0 else 0.0
        rv_se  = bootstrap_se(rv_df,"weight_rv",qcol,resp,500)
        lw = pd.to_numeric(lv_df["weight_lv"],errors="coerce").fillna(0)
        lq = lv_df[qcol].astype(str).str.strip()
        lv_pct = lw[lq==resp].sum()/lw.sum()*100 if lw.sum()>0 else 0.0
        lv_se  = bootstrap_se(lv_df,"weight_lv",qcol,resp,500)
        print(f"  {qcol.replace('_',' ')[:25]:<25}  {resp[:30]:<30}  {rv_pct:>5.1f}%  {rv_se:>5.2f}  {lv_pct:>5.1f}%  {lv_se:>5.2f}")
    print(f"{'='*60}")

    # ── Export Excel ───────────────────────────────────────────────────────────
    global GROUP_ORDER, _CURRENT_GROUP_ORDER_REF

    GROUP_ORDER = [
        ("Race (W College/No College)",
         "RaceEdu",
         ["White No College","White College","Hispanic","Black","Asian / Other"]),
        ("Age",           "Age",         ["18-29","30-44","45-64","65+"]),
        ("Gender",        "Gender",       ["Male","Female"]),
        ("Age × Gender",  "AgeGender",
         ["18-29_Male","18-29_Female","30-44_Male","30-44_Female",
          "45-64_Male","45-64_Female","65+_Male","65+_Female"]),
        ("Education",     "Education4",
         ["High school or less","Some college/assoc. degree",
          "College graduate","Postgraduate study"]),
        ("Income",        "Income",
         ["$0–$25k","$25–$50k","$50–$75k","$75–$100k","$100–$150k","$150–$200k","$200k+"]),
        ("Region",        "Region",       NATIONAL_REGIONS),
        ("Party",         "Party",        ["Republican","Democrat","Independent"]),
        ("Party Detail",  "Party_Detailed", PARTY_DETAILED_ORDER_DISPLAY),
        ("2024 Vote",     "Vote2024_Bucket",
         ["Donald Trump","Kamala Harris","Third party","Did not vote"]),
        ("Trump Approval","Q14_TrumpApprove", APPROVE_RESPONSE_ORDER),
        ("Vote History",  "VoteHistory",
         ["Consistent voter","Occasional voter","New / non-voter"]),
        ("Groyper (Q10)",  "Groyper_Bucket",  ["Yes", "No"]),
    ]
    _CURRENT_GROUP_ORDER_REF = GROUP_ORDER

    print(f"\n{'='*60}\n  EXPORTING EXCEL: {OUTPUT_FILE}\n{'='*60}")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        _init = writer.book.create_sheet("_init_")

        print("\n  Building Weight Diagnostics sheet...")
        write_diagnostics_sheet(writer, diag_rv, diag_lv,
                                 balance_rv, balance_lv, icc_results, moe_df)

        print("  Writing raw data sheets...")
        rv_df.to_excel(writer, sheet_name="RV_Weighted", index=False)
        lv_df.to_excel(writer, sheet_name="LV_Weighted", index=False)
        print("  ✅ RV_Weighted, LV_Weighted done")
        for sname,dfi,wcol in [
            ("RV_Tabbook", rv_df, "weight_rv"),
            ("LV_Tabbook", lv_df, "weight_lv"),
        ]:
            t1 = _time.time()
            print(f"\n  Building {sname}...")
            write_tabbook_sheet(writer,sname,dfi.copy(),wcol,q7_rank_cols,resolved_matrix)
            print(f"  ✅ {sname} done in {_time.time()-t1:.1f}s")
        print("\n  Building Electorate_Composition...")
        build_electorate_sheet(writer,"Electorate_Composition",rv_df,lv_df)
        print("  ✅ Electorate_Composition done")

        if "_init_" in writer.book.sheetnames:
            del writer.book["_init_"]
        desired  = ["Weight Diagnostics", "Electorate_Composition",
                    "RV_Tabbook", "LV_Tabbook", "RV_Weighted", "LV_Weighted"]
        existing = writer.book.sheetnames
        ordered  = [s for s in desired if s in existing] + [s for s in existing if s not in desired]
        writer.book._sheets.sort(
            key=lambda ws: ordered.index(ws.title) if ws.title in ordered else 9999
        )

    print(f"\n  ✅ Excel saved: {OUTPUT_FILE}")

    # ── Terminal toplines ──────────────────────────────────────────────────────
    buf = io.StringIO(); orig = sys.stdout
    sys.stdout = type("T",(object,),{
        "write": lambda s,d: [orig.write(d), buf.write(d)],
        "flush": lambda s:   [orig.flush(),  buf.flush()],
    })()
    print_all_questions(rv_df, lv_df, q7_rank_cols)
    sys.stdout = orig
    with open(TOPLINES_FILE,"w",encoding="utf-8") as f:
        f.write(f"SURVEY TOPLINES — {SURVEY_NAME}\n")
        f.write("="*80+"\n\n")
        f.write(buf.getvalue())
    print(f"  ✅ Toplines saved: {TOPLINES_FILE}")

    elapsed = _time.time() - t0
    print(f"\n{'='*60}")
    print(f"  ✅  DONE in {elapsed:.1f}s — {SURVEY_NAME}")
    print(f"  RV: {len(rv_df):,}  |  LV: {len(lv_df):,}")
    print(f"  Weight cap: {WEIGHT_CAP}  |  DEFF(RV)={deff_final:.3f}  |  DEFF(LV)={deff_lv:.3f}")
    if icc_results.get("ICC"):
        print(f"  ICC={icc_results['ICC']}  DEFF_cluster={icc_results['DEFF_cluster']}")
    print(f"  Outputs:  {OUTPUT_FILE}  |  {TOPLINES_FILE}")
    print(f"{'='*60}\n")

    print_deff_summary("Final RV", rv_df["weight_rv"])
    print_deff_summary("Final LV", lv_df["weight_lv"])

    return rv_df, lv_df

# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    CSV_PATH = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    rv_df, lv_df = main(csv_path=CSV_PATH)